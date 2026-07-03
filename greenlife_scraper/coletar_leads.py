"""
coletar_leads.py — Coleta leads CRM Pacto (versão para GitHub Actions).
Tokens carregados via env var PACTO_TOKENS (JSON).
"""

import urllib.request, urllib.error, json, ssl, time, calendar, os
from datetime import datetime, timezone, timedelta, date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ctx      = ssl.create_default_context()
BRT      = timezone(timedelta(hours=-3))
BASE_DIR = Path(__file__).parent
EXCEL    = BASE_DIR / "historico_leads.xlsx"
JSON_OUT = BASE_DIR / "leads_dados.json"

HISTORICO_INICIO = (2024, 1)

# Tokens carregados do env var PACTO_TOKENS (JSON: {"Aldeota": "token...", ...})
_tokens = json.loads(os.environ.get("PACTO_TOKENS", "{}"))

UNIDADES = [
    ("Aldeota",          _tokens.get("Aldeota",          ""), 2, "Regiao 1"),
    ("CT",               _tokens.get("CT",               ""), 1, "Regiao 1"),
    ("Guararapes",       _tokens.get("Guararapes",       ""), 1, "Regiao 1"),
    ("Personal",         _tokens.get("Personal",         ""), 1, "Regiao 1"),
    ("Riomar",           _tokens.get("Riomar",           ""), 3, "Regiao 1"),
    ("Rui Barbosa",      _tokens.get("Rui Barbosa",      ""), 1, "Regiao 1"),
    ("Shopping Aldeota", _tokens.get("Shopping Aldeota", ""), 1, "Regiao 1"),
    ("Passare",          _tokens.get("Passare",          ""), 1, "Regiao 2"),
    ("Fatima",           _tokens.get("Fatima",           ""), 1, "Regiao 2"),
    ("Kennedy",          _tokens.get("Kennedy",          ""), 1, "Regiao 2"),
    ("Maraponga",        _tokens.get("Maraponga",        ""), 1, "Regiao 2"),
    ("Montese",          _tokens.get("Montese",          ""), 1, "Regiao 2"),
    ("Parquelandia",     _tokens.get("Parquelandia",     ""), 1, "Regiao 2"),
    ("Joquei",           _tokens.get("Joquei",           ""), 1, "Regiao 2"),
    ("Cambeba",          _tokens.get("Cambeba",          ""), 1, "Regiao 3"),
    ("Caucaia",          _tokens.get("Caucaia",          ""), 1, "Regiao 3"),
    ("Eusebio",          _tokens.get("Eusebio",          ""), 1, "Regiao 3"),
    ("Maranguape",       _tokens.get("Maranguape",       ""), 1, "Regiao 3"),
    ("Messejana",        _tokens.get("Messejana",        ""), 1, "Regiao 3"),
    ("Sul",              _tokens.get("Sul",              ""), 1, "Regiao 3"),
    ("Barra Funda",      _tokens.get("Barra Funda",      ""), 1, "Sao Paulo"),
    ("Tatuape",          _tokens.get("Tatuape",          ""), 1, "Sao Paulo"),
    ("Moema",            _tokens.get("Moema",            ""), 1, "Sao Paulo"),
    ("CT norte",         _tokens.get("CT norte",         ""), 1, "Regiao 2"),
]


def ts_ms(year, month, day):
    return int(datetime(year, month, day, 0, 0, 0, tzinfo=BRT).timestamp() * 1000)

def meses_para_coletar():
    hoje = date.today()
    ano, mes = HISTORICO_INICIO
    meses = []
    while (ano, mes) <= (hoje.year, hoje.month):
        _, ultimo = calendar.monthrange(ano, mes)
        fim_dia = hoje.day if (ano, mes) == (hoje.year, hoje.month) else ultimo
        meses.append((ano, mes, ts_ms(ano, mes, 1), ts_ms(ano, mes, fim_dia)))
        mes += 1
        if mes > 12:
            mes = 1; ano += 1
    return meses

def _chamar_api(req, tentativas=4):
    for tentativa in range(1, tentativas + 1):
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=20) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code == 429:
                espera = 60 * tentativa
                print(f"\n    [429] Rate limit. Aguardando {espera}s ({tentativa}/{tentativas-1})...", end=" ", flush=True)
                time.sleep(espera)
            else:
                raise
    raise RuntimeError(f"API retornou 429 em todas as tentativas.")

def get_usuarios(api_key, empresa_id):
    req = urllib.request.Request(
        "https://apigw.pactosolucoes.com.br/bi-crm/usuarios",
        headers={"Authorization": f"Bearer {api_key}", "empresaId": str(empresa_id)},
    )
    return [u["codigo"] for u in _chamar_api(req).get("content", [])]

def get_leads(api_key, empresa_id, usuarios, ini, fim):
    body = json.dumps({"indicador": "RESULTADO", "dataInicio": ini, "dataFim": fim,
                       "codigosUsuarios": usuarios}).encode()
    req = urllib.request.Request(
        "https://apigw.pactosolucoes.com.br/bi-crm", data=body,
        headers={"Authorization": f"Bearer {api_key}", "empresaId": str(empresa_id),
                 "Content-Type": "application/json"},
        method="POST",
    )
    res = _chamar_api(req)["content"]["resultado"]
    cl = next((x for x in res if x["identificadorMeta"] == "CL"), {})
    return int(cl.get("meta", 0)), int(cl.get("metaAtingida", 0))

def carregar_historico():
    if not EXCEL.exists():
        return {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Histórico"]
    hist = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            hist[(str(row[0]), int(row[2]), int(row[3]))] = int(row[4] or 0)
    return hist

def salvar_excel(dados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"
    ws.append(["Unidade", "Região", "Ano", "Mês", "Leads", "Coletado em"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B3A2F")
        cell.alignment = Alignment(horizontal="center")
    for (unidade, ano, mes), (regiao, leads, coletado) in sorted(dados.items()):
        ws.append([unidade, regiao, ano, mes, leads, coletado])
    for col, w in zip("ABCDEF", [22, 12, 8, 6, 10, 20]):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet("Resumo Mensal")
    _gerar_aba_resumo(ws2, dados)
    wb.save(EXCEL)

def _gerar_aba_resumo(ws, dados):
    meses_uniq = sorted({(ano, mes) for _, ano, mes in dados})
    regioes = ["Regiao 1", "Regiao 2", "Regiao 3", "Sao Paulo", "TOTAL"]
    ws.append(["Região"] + [f"{m:02d}/{a}" for a, m in meses_uniq])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B3A2F")
        cell.alignment = Alignment(horizontal="center")
    for regiao in regioes:
        linha = [regiao]
        for ano, mes in meses_uniq:
            if regiao == "TOTAL":
                total = sum(v for (_, a, m), (_, v, _) in dados.items() if a == ano and m == mes)
            else:
                total = sum(v for (_, a, m), (r, v, _) in dados.items() if a == ano and m == mes and r == regiao)
            linha.append(total)
        ws.append(linha)
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(meses_uniq) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 10

def exportar_json(dados, conv_dados=None):
    if conv_dados is None:
        conv_dados = {}
    hoje = date.today()
    meses_uniq = sorted({(ano, mes) for _, ano, mes in dados})
    labels = [f"{m:02d}/{a}" for a, m in meses_uniq]
    por_unidade = {}
    for (unidade, ano, mes), (regiao, leads, _) in dados.items():
        if unidade not in por_unidade:
            por_unidade[unidade] = {"regiao": regiao, "meses": {}, "conversao": {}}
        chave = f"{ano}-{mes:02d}"
        por_unidade[unidade]["meses"][chave] = leads
        conv = conv_dados.get((unidade, ano, mes))
        if conv is not None:
            por_unidade[unidade]["conversao"][chave] = conv
    regioes = {}
    for (_, ano, mes), (regiao, leads, _) in dados.items():
        key = f"{ano}-{mes:02d}"
        regioes.setdefault(regiao, {})[key] = regioes.get(regiao, {}).get(key, 0) + leads
    totais_mes = {}
    for (_, ano, mes), (_, leads, _) in dados.items():
        key = f"{ano}-{mes:02d}"
        totais_mes[key] = totais_mes.get(key, 0) + leads
    resultado = {
        "gerado_em":  datetime.now(BRT).strftime("%d/%m/%Y %H:%M"),
        "mes_atual":  f"{hoje.month:02d}/{hoje.year}",
        "labels":     labels,
        "unidades":   por_unidade,
        "regioes":    regioes,
        "totais_mes": totais_mes,
    }
    JSON_OUT.parent.mkdir(exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON exportado -> {JSON_OUT}")

def main():
    if not any(t for _, t, _, _ in UNIDADES):
        print("[ERRO] PACTO_TOKENS não definido. Configure o GitHub Secret.")
        raise SystemExit(1)

    hoje  = date.today()
    meses = meses_para_coletar()
    hist  = carregar_historico()
    dados = {k: ("", v, "") for k, v in hist.items()}
    agora = datetime.now(BRT).strftime("%Y-%m-%d %H:%M")
    conv_dados = {}

    total_unidades = len(UNIDADES)
    for idx, (nome, api_key, empresa_id, regiao) in enumerate(UNIDADES, 1):
        if not api_key:
            print(f"  [{idx}/{total_unidades}] {nome} — token ausente, pulando")
            continue
        pendentes = [
            (ano, mes, ini, fim) for ano, mes, ini, fim in meses
            if (nome, ano, mes) not in hist or (ano == hoje.year and mes == hoje.month)
        ]
        if not pendentes:
            print(f"  [{idx}/{total_unidades}] {nome} — sem novidades, pulando")
            for ano, mes, _, _ in meses:
                if (nome, ano, mes) in hist:
                    dados[(nome, ano, mes)] = (regiao, hist[(nome, ano, mes)], agora)
            continue

        print(f"  [{idx}/{total_unidades}] {nome} ({len(pendentes)} meses)...", end=" ", flush=True)
        try:
            usuarios = get_usuarios(api_key, empresa_id)
            for ano, mes, ini, fim in pendentes:
                v, conv = get_leads(api_key, empresa_id, usuarios, ini, fim)
                dados[(nome, ano, mes)] = (regiao, v, agora)
                conv_dados[(nome, ano, mes)] = conv
                time.sleep(0.8)
            for ano, mes, _, _ in meses:
                if (nome, ano, mes) in hist and (nome, ano, mes) not in dados:
                    dados[(nome, ano, mes)] = (regiao, hist[(nome, ano, mes)], agora)
            print("OK")
        except Exception as e:
            print(f"ERRO: {e}")
            for ano, mes, _, _ in meses:
                if (nome, ano, mes) in hist:
                    dados[(nome, ano, mes)] = (regiao, hist[(nome, ano, mes)], agora)

        salvar_excel(dados)
        hist = carregar_historico()
        time.sleep(2)

    exportar_json(dados, conv_dados)
    print(f"\n[CONCLUIDO] {sum(v for _, (_, v, _) in dados.items() if v)} leads total.")

if __name__ == "__main__":
    main()
