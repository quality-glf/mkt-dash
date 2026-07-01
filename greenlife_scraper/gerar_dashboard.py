# -*- coding: utf-8 -*-
"""Lê os Excel e gera dashboard.html com abas: Visão Geral, Instagram, Parceiros, Leads Pacto."""
import openpyxl
import json as _json
import calendar as _cal
from pathlib import Path
from datetime import datetime, date

INSTAGRAM_DIR = Path(__file__).parent.parent / "instagram-metrics"
LEADS_JSON    = Path(__file__).parent / "leads_dados.json"
METAS_JSON    = Path(__file__).parent / "metas.json"

MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
            "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

PERFIS_ORDEM = ["Academias", "São Paulo", "Family", "Personal", "CT"]

CORES_PERFIS = {
    "Academias":  "#2E7D32",
    "São Paulo":  "#1565C0",
    "Family":     "#F57C00",
    "Personal":   "#6A1B9A",
    "CT":         "#00838F",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def fmt(v):
    if isinstance(v, (int, float)):
        return f"{int(v):,}".replace(",", ".")
    return str(v) if v is not None else "—"


def fmt_mes(mk):
    """'06/2026' → 'Junho/2026'"""
    try:
        m, y = mk.split("/")
        return f"{MESES_PT[int(m)-1]}/{y}"
    except Exception:
        return mk


def dias_uteis_no_mes(ano, mes):
    _, ultimo = _cal.monthrange(ano, mes)
    return sum(1 for d in range(1, ultimo + 1) if date(ano, mes, d).weekday() < 5)


def dias_uteis_ate(ano, mes, ate_dia):
    ate_dia = max(1, min(ate_dia, _cal.monthrange(ano, mes)[1]))
    return sum(1 for d in range(1, ate_dia + 1) if date(ano, mes, d).weekday() < 5)


def _mes_key(data_fim_str):
    try:
        parts = data_fim_str.split("/")
        return f"{parts[1]}/{parts[2]}"
    except Exception:
        return "?"


def _mes_sort(mes_key):
    try:
        m, y = mes_key.split("/")
        return (int(y), int(m))
    except Exception:
        return (0, 0)


# ── Metas ─────────────────────────────────────────────────────────────────────

METAS_XLSX = Path(__file__).parent / "metas.xlsx"
MESES_PT_ABR = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

def _label_to_key(label):
    """'Jun/2026' -> '2026-06'"""
    try:
        mes_str, ano = label.strip().split("/")
        mes = MESES_PT_ABR.index(mes_str[:3].capitalize()) + 1
        return f"{ano}-{mes:02d}"
    except Exception:
        return None

def _key_to_label(k):
    """'2026-06' -> 'Jun/2026'"""
    try:
        return f"{MESES_PT_ABR[int(k[5:])-1]}/{k[:4]}"
    except Exception:
        return k

def carregar_metas():
    """Lê metas do Excel. Fallback para JSON se Excel não existir."""
    if METAS_XLSX.exists():
        return _carregar_metas_excel()
    if METAS_JSON.exists():
        with open(METAS_JSON, encoding="utf-8") as f:
            return _json.load(f)
    return {"leads": {}, "seguidores": {}}

def _carregar_metas_excel():
    wb = openpyxl.load_workbook(METAS_XLSX, data_only=True)
    resultado = {"leads": {}, "seguidores": {}, "interacoes": {}}

    def _ler_aba_perfis(ws):
        headers = [c.value for c in ws[1]]
        mes_keys = [_label_to_key(str(h)) if h else None for h in headers[1:]]
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            perfil = row[0]
            if not perfil or str(perfil).startswith("Adicione"):
                continue
            p_metas = {}
            for i, mk in enumerate(mes_keys):
                val = row[i + 1]
                if mk and val is not None:
                    try:
                        p_metas[mk] = int(val)
                    except (ValueError, TypeError):
                        pass
            if p_metas:
                out[str(perfil)] = p_metas
        return out

    # Aba Leads
    if "Leads por Unidade" in wb.sheetnames:
        ws = wb["Leads por Unidade"]
        headers = [c.value for c in ws[1]]
        mes_keys = [_label_to_key(str(h)) if h else None for h in headers[2:]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            if not nome or str(nome).startswith("Edite"):
                continue
            u_metas = {}
            for i, mk in enumerate(mes_keys):
                val = row[i + 2]
                if mk and val is not None:
                    try:
                        u_metas[mk] = int(val)
                    except (ValueError, TypeError):
                        pass
            if u_metas:
                resultado["leads"][str(nome)] = u_metas

    if "Seguidores Instagram" in wb.sheetnames:
        resultado["seguidores"] = _ler_aba_perfis(wb["Seguidores Instagram"])

    if "Interações Instagram" in wb.sheetnames:
        resultado["interacoes"] = _ler_aba_perfis(wb["Interações Instagram"])

    return resultado

def _criar_metas_excel(leads_metas, seg_metas):
    """Cria metas.xlsx do zero com os valores calculados."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COR_H = "1B3A2F"
    COR_S = "6A1B9A"
    thin  = Side(style="thin", color="D1D5DB")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads por Unidade"
    ws.freeze_panes = "C2"

    todos_meses = sorted(
        {mk for u in leads_metas.values() for mk in u},
        key=lambda k: (k[:4], k[5:])
    )
    headers = ["Unidade", "Região"] + [_key_to_label(m) for m in todos_meses]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor=COR_H)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda
    ws.row_dimensions[1].height = 30

    ORDEM = ["Aldeota","CT","Guararapes","Personal","Riomar","Rui Barbosa","Shopping Aldeota",
             "Passare","Fatima","Kennedy","Maraponga","Montese","Parquelandia","Joquei",
             "Cambeba","Caucaia","Eusebio","Maranguape","Messejana","Sul",
             "Barra Funda","Tatuape","Moema","CT norte"]
    REGIOES = {"Aldeota":"Região 1","CT":"Região 1","Guararapes":"Região 1","Personal":"Região 1",
               "Riomar":"Região 1","Rui Barbosa":"Região 1","Shopping Aldeota":"Região 1",
               "Passare":"Região 2","Fatima":"Região 2","Kennedy":"Região 2","Maraponga":"Região 2",
               "Montese":"Região 2","Parquelandia":"Região 2","Joquei":"Região 2",
               "Cambeba":"Região 3","Caucaia":"Região 3","Eusebio":"Região 3",
               "Maranguape":"Região 3","Messejana":"Região 3","Sul":"Região 3",
               "Barra Funda":"São Paulo","Tatuape":"São Paulo","Moema":"São Paulo","CT norte":"Região 2"}

    for row_i, nome in enumerate(ORDEM, 2):
        u_metas = leads_metas.get(nome, {})
        zebra   = row_i % 2 == 0
        vals    = [nome, REGIOES.get(nome, "")] + [u_metas.get(mk) for mk in todos_meses]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = borda
            if zebra:
                c.fill = PatternFill("solid", fgColor="F0FDF4")
            if col == 1:
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = Font(bold=True)
            if col == 2:
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = Font(color="6B7280")
            if val and col > 2:
                c.number_format = "#,##0"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10

    inst = ws.cell(row=len(ORDEM) + 3, column=1,
                   value="Edite os valores diretamente. Deixe vazio para sem meta. Salve e re-execute gerar_dashboard.py.")
    inst.font = Font(italic=True, color="9CA3AF", size=10)

    # Aba seguidores
    ws2 = wb.create_sheet("Seguidores Instagram")
    ws2.freeze_panes = "B2"
    PERFIS = ["Academias", "São Paulo", "Family", "Personal"]
    meses_seg = sorted({mk for p in seg_metas.values() for mk in p}, key=lambda k: (k[:4], k[5:])) if seg_metas else []
    headers2 = ["Perfil"] + ([_key_to_label(m) for m in meses_seg] if meses_seg else [])
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor=COR_S)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borda
    ws2.row_dimensions[1].height = 30
    for row_i, perfil in enumerate(PERFIS, 2):
        p_metas = seg_metas.get(perfil, {})
        vals2 = [perfil] + [p_metas.get(mk) for mk in meses_seg]
        for col, val in enumerate(vals2, 1):
            c = ws2.cell(row=row_i, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = borda
            if row_i % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F3E5F5")
            if col == 1:
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = Font(bold=True)
            if val and col > 1:
                c.number_format = "#,##0"
    ws2.column_dimensions["A"].width = 16
    for i in range(2, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 12
    inst2 = ws2.cell(row=7, column=1,
                     value="Adicione colunas no formato Mmm/AAAA (ex: Jul/2026) conforme necessário.")
    inst2.font = Font(italic=True, color="9CA3AF", size=10)

    wb.save(METAS_XLSX)


def inicializar_metas(leads_data):
    """Adiciona meses faltando no metas.xlsx com 5% acima do mês anterior."""
    if not METAS_XLSX.exists():
        # Cria do zero com valores calculados a partir do histórico real
        leads_metas = {}
        unidades = leads_data.get("unidades", {})
        for nome, info in unidades.items():
            meses_hist = sorted(info.get("meses", {}).items())
            u_metas = {}
            prev_val = None
            for mes_key, valor in meses_hist:
                if prev_val is not None:
                    u_metas[mes_key] = round(prev_val * 1.05)
                prev_val = valor
            # próximo mês
            if meses_hist:
                lk, lv = meses_hist[-1]
                ano_l, mes_l = int(lk[:4]), int(lk[5:])
                prox = f"{ano_l}-{mes_l+1:02d}" if mes_l < 12 else f"{ano_l+1}-01"
                u_metas[prox] = round(lv * 1.05)
            leads_metas[nome] = u_metas
        _criar_metas_excel(leads_metas, {})
        return {"leads": leads_metas, "seguidores": {}}

    metas_atuais = _carregar_metas_excel()
    leads_metas  = metas_atuais.get("leads", {})
    wb = openpyxl.load_workbook(METAS_XLSX)

    if "Leads por Unidade" not in wb.sheetnames:
        return metas_atuais

    ws = wb["Leads por Unidade"]
    headers = [c.value for c in ws[1]]
    mes_labels_existentes = set(str(h) for h in headers[2:] if h)

    # Descobre meses novos que ainda não têm coluna
    unidades_data = leads_data.get("unidades", {})
    todos_meses_leads = sorted(
        {mk for u in unidades_data.values() for mk in u.get("meses", {})},
        key=lambda k: (k[:4], k[5:])
    )
    novos = [mk for mk in todos_meses_leads
             if _key_to_label(mk) not in mes_labels_existentes]

    # Adiciona próximo mês futuro
    if todos_meses_leads:
        lk = todos_meses_leads[-1]
        ano_l, mes_l = int(lk[:4]), int(lk[5:])
        prox = f"{ano_l}-{mes_l+1:02d}" if mes_l < 12 else f"{ano_l+1}-01"
        if _key_to_label(prox) not in mes_labels_existentes:
            novos.append(prox)

    if novos:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        COR_H = "1B3A2F"
        thin  = Side(style="thin", color="D1D5DB")
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)

        for mk in novos:
            col_idx = ws.max_column + 1
            label   = _key_to_label(mk)
            c = ws.cell(row=1, column=col_idx, value=label)
            c.font      = Font(bold=True, color="FFFFFF", size=11)
            c.fill      = PatternFill("solid", fgColor=COR_H)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = borda
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

            # Preenche linhas com 5% do mês anterior
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                nome = row[0].value
                if not nome or str(nome).startswith("Edite"):
                    continue
                u_metas = leads_metas.get(str(nome), {})
                # encontra valor do mês anterior para esta unidade
                mes_keys_ord = sorted(u_metas.keys(), key=lambda k: (k[:4], k[5:]))
                prev_val = u_metas.get(mes_keys_ord[-1]) if mes_keys_ord else None
                if prev_val:
                    cell = ws.cell(row=row[0].row, column=col_idx, value=round(prev_val * 1.05))
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = borda

        wb.save(METAS_XLSX)

    return _carregar_metas_excel()


# ── Carregamento de dados ─────────────────────────────────────────────────────

def carregar_instagram():
    seg_rows, int_rows, det_rows = [], [], []

    arq = INSTAGRAM_DIR / "base_seguidores.xlsx"
    if arq.exists():
        ws = openpyxl.load_workbook(arq).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                seg_rows.append({"perfil": row[1], "data_fim": row[3],
                                  "ganhos": row[5], "perdidos": row[6], "crescimento": row[7],
                                  "total": row[4], "semana": row[0], "data_ini": row[2]})

    arq = INSTAGRAM_DIR / "base_interacoes_total.xlsx"
    if arq.exists():
        ws = openpyxl.load_workbook(arq).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                int_rows.append({"perfil": row[1], "data_fim": row[3], "total_int": row[4]})

    arq = INSTAGRAM_DIR / "base_interacoes_detalhadas.xlsx"
    if arq.exists():
        ws = openpyxl.load_workbook(arq).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                det_rows.append({"perfil": row[1], "data_fim": row[3],
                                  "curtidas": row[4], "comentarios": row[5],
                                  "compartilhamentos": row[6], "salvamentos": row[7], "posts": row[8]})

    seg_latest, int_latest, det_latest = {}, {}, {}
    for r in seg_rows:
        seg_latest[r["perfil"]] = r
    for r in int_rows:
        int_latest[r["perfil"]] = r
    for r in det_rows:
        det_latest[r["perfil"]] = r

    perfis = []
    for nome in PERFIS_ORDEM:
        s = seg_latest.get(nome, {})
        it = int_latest.get(nome, {})
        d = det_latest.get(nome, {})
        perfis.append({
            "nome": nome, "semana": s.get("semana", "—"),
            "data_ini": s.get("data_ini", "—"), "data_fim": s.get("data_fim", "—"),
            "total": s.get("total"), "ganhos": s.get("ganhos"),
            "perdidos": s.get("perdidos"), "crescimento": s.get("crescimento"),
            "total_int": it.get("total_int"), "curtidas": d.get("curtidas"),
            "comentarios": d.get("comentarios"), "compartilhamentos": d.get("compartilhamentos"),
            "salvamentos": d.get("salvamentos"), "posts": d.get("posts"),
        })

    indicadores = ["total_seguidores", "ganhos", "perdidos", "crescimento",
                   "total_int", "curtidas", "comentarios",
                   "compartilhamentos", "salvamentos", "posts"]
    hist = {ind: {p: {} for p in PERFIS_ORDEM} for ind in indicadores}
    meses_set = set()

    for r in seg_rows:
        mk = _mes_key(r["data_fim"]); meses_set.add(mk)
        p = r["perfil"]
        for ind in ["ganhos", "perdidos", "crescimento"]:
            v = r.get(ind)
            if v is not None and p in hist[ind]:
                hist[ind][p][mk] = v
        if r.get("total") is not None and p in hist["total_seguidores"]:
            hist["total_seguidores"][p][mk] = r["total"]

    for r in int_rows:
        mk = _mes_key(r["data_fim"]); meses_set.add(mk)
        p = r["perfil"]
        if p in hist["total_int"]:
            v = r.get("total_int")
            if v is not None:
                hist["total_int"][p][mk] = v

    for r in det_rows:
        mk = _mes_key(r["data_fim"]); meses_set.add(mk)
        p = r["perfil"]
        for ind in ["curtidas", "comentarios", "compartilhamentos", "salvamentos", "posts"]:
            v = r.get(ind)
            if v is not None and p in hist[ind]:
                hist[ind][p][mk] = v

    # Histórico de total de seguidores por mês/perfil
    seg_total_hist = {}
    for r in seg_rows:
        mk = _mes_key(r["data_fim"])
        if r["total"] is not None:
            seg_total_hist.setdefault(mk, {})[r["perfil"]] = r["total"]

    meses = sorted(meses_set, key=_mes_sort)
    return perfis, hist, meses, seg_total_hist


def carregar_parceiros(arquivo_excel):
    wb = openpyxl.load_workbook(arquivo_excel)

    ws_r = wb["Resumo Diario"]
    linhas = [r for r in ws_r.iter_rows(min_row=2, values_only=True) if r[0]]
    ultima = linhas[-1] if linhas else ("—",) * 7

    resumo = {
        "data": ultima[0] or "—", "total_promocoes": ultima[1] or "—",
        "cupons_gerados": ultima[2] or "—", "cupons_utilizados": ultima[3] or "—",
        "total_parceiros": ultima[4] or "—", "parc_com_usuario": ultima[5] or "—",
        "parc_sem_usuario": ultima[6] or "—",
    }

    ws_parc = wb["Parceiros"]
    parceiros = {}
    for row in ws_parc.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            parceiros[str(row[0])] = row[1]

    ws_promo = wb["Promocoes"]
    cupons_por_parceiro = {}
    for row in ws_promo.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[9]:
            continue
        try:
            pid = str(int(float(str(row[9]))))
            gerado = int(float(str(row[5]))) if row[5] and str(row[5]).replace(".", "").isdigit() else 0
            usado  = int(float(str(row[6]))) if row[6] and str(row[6]).replace(".", "").isdigit() else 0
            if pid not in cupons_por_parceiro:
                cupons_por_parceiro[pid] = {"gerados": 0, "utilizados": 0}
            cupons_por_parceiro[pid]["gerados"] += gerado
            cupons_por_parceiro[pid]["utilizados"] += usado
        except Exception:
            pass

    ranking = [
        {"nome": parceiros.get(pid, f"Parceiro {pid}"), "gerados": v["gerados"],
         "utilizados": v["utilizados"],
         "taxa": round(v["utilizados"] / v["gerados"] * 100, 1) if v["gerados"] else 0}
        for pid, v in cupons_por_parceiro.items() if v["utilizados"] > 0
    ]
    ranking.sort(key=lambda x: -x["utilizados"])

    hist_mensal = {}
    for r in linhas:
        if not r[0] or not r[2]:
            continue
        try:
            data_str = str(r[0])
            partes = data_str[:10].split("-")
            mk = f"{partes[1]}/{partes[0]}"
            g = int(r[2]) if r[2] else 0
            u = int(r[3]) if r[3] else 0
            if mk not in hist_mensal:
                hist_mensal[mk] = {"gerados": g, "utilizados": u}
            else:
                hist_mensal[mk]["gerados"]    = max(hist_mensal[mk]["gerados"], g)
                hist_mensal[mk]["utilizados"] = max(hist_mensal[mk]["utilizados"], u)
        except Exception:
            pass

    ws_st = wb["Serie Temporal"]
    serie_mensal = {}
    for row in ws_st.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        try:
            partes = str(row[0]).split("/")
            if len(partes) != 3 or len(partes[2]) != 4:
                continue
            mk = f"{partes[1]}/{partes[2]}"
            g = int(row[1]) if row[1] else 0
            u = int(row[2]) if row[2] else 0
            if mk not in serie_mensal:
                serie_mensal[mk] = {"gerados": 0, "utilizados": 0}
            serie_mensal[mk]["gerados"]    += g
            serie_mensal[mk]["utilizados"] += u
        except Exception:
            pass

    todos_meses = set(hist_mensal) | set(serie_mensal)
    mensal_final = {}
    for mk in todos_meses:
        mensal_final[mk] = serie_mensal[mk] if mk in serie_mensal else hist_mensal[mk]

    mensal_sorted = sorted(mensal_final.items(),
                           key=lambda x: (x[0][3:], x[0][:2]), reverse=True)
    return resumo, ranking, mensal_sorted


def carregar_leads():
    if not LEADS_JSON.exists():
        return None
    with open(LEADS_JSON, encoding="utf-8") as f:
        return _json.load(f)


# ── Geração do dashboard ──────────────────────────────────────────────────────

def gerar_dashboard(arquivo_excel="dados_greenlife.xlsx", saida="dashboard.html"):
    resumo, ranking, hist_mensal_sorted = carregar_parceiros(arquivo_excel)
    ig, ig_hist, ig_meses, seg_total_hist = carregar_instagram()
    leads = carregar_leads()

    semana_ig  = ig[0]["semana"] if ig else "—"
    periodo_ig = f"{ig[0]['data_ini']} → {ig[0]['data_fim']}" if ig else "—"
    total_seguidores  = sum(p["total"] or 0 for p in ig)
    total_ganhos      = sum(p["ganhos"] or 0 for p in ig)
    total_interacoes  = sum(p["total_int"] or 0 for p in ig)

    hoje = date.today()
    mes_atual = hoje.strftime("%m/%Y")

    # ── Metas ─────────────────────────────────────────────────────
    if leads:
        metas_data  = inicializar_metas(leads)
    else:
        metas_data  = carregar_metas()
    metas_leads = metas_data.get("leads", {})
    metas_seg   = metas_data.get("seguidores", {})

    # ── LEADS: variáveis derivadas ─────────────────────────────────
    if leads:
        totais_mes = leads.get("totais_mes", {})
        unidades   = leads.get("unidades", {})
        regioes    = leads.get("regioes", {})
        gerado_em  = leads.get("gerado_em", "—")
        mes_atual_key = leads.get("mes_atual", "")

        if mes_atual_key:
            partes = mes_atual_key.split("/")
            chave_atual = f"{partes[1]}-{partes[0]}"
        else:
            chave_atual = ""

        leads_mes_atual = totais_mes.get(chave_atual, 0)

        # Metas de leads — séries completas
        labels_todos = leads.get("labels", [])
        keys_todos   = [f"{l[3:]}-{l[:2]}" for l in labels_todos]

        def serie_rede():
            return [totais_mes.get(k, 0) for k in keys_todos]

        def serie_regiao(nome):
            d = regioes.get(nome, {})
            return [d.get(k, 0) for k in keys_todos]

        def serie_unidade(nome):
            u = unidades.get(nome, {}).get("meses", {})
            return [u.get(k, 0) for k in keys_todos]

        # Metas por mês
        meta_rede_por_mes = {}
        for nome_u, u_m in metas_leads.items():
            for mk, v in u_m.items():
                meta_rede_por_mes[mk] = meta_rede_por_mes.get(mk, 0) + v

        meta_reg_por_mes = {"Regiao 1": {}, "Regiao 2": {}, "Regiao 3": {}, "Sao Paulo": {}}
        for nome_u, u_m in metas_leads.items():
            reg = unidades.get(nome_u, {}).get("regiao", "")
            if reg in meta_reg_por_mes:
                for mk, v in u_m.items():
                    meta_reg_por_mes[reg][mk] = meta_reg_por_mes[reg].get(mk, 0) + v

        def meta_serie_rede():
            return [meta_rede_por_mes.get(k) for k in keys_todos]

        def meta_serie_regiao(nome):
            d = meta_reg_por_mes.get(nome, {})
            return [d.get(k) for k in keys_todos]

        def meta_serie_unidade(nome):
            u = metas_leads.get(nome, {})
            return [u.get(k) for k in keys_todos]

        # Meta do dia para o card
        meta_mes_rede = meta_rede_por_mes.get(chave_atual, 0)
        if meta_mes_rede > 0:
            du_mes      = dias_uteis_no_mes(hoje.year, hoje.month)
            du_passados = dias_uteis_ate(hoje.year, hoje.month, hoje.day)
            meta_ate_hoje = round(meta_mes_rede * du_passados / du_mes) if du_mes else 0
            vs_meta       = leads_mes_atual - meta_ate_hoje
            vs_meta_str   = f"{'+'if vs_meta>=0 else ''}{fmt(vs_meta)}"
            vs_meta_cor   = "#2E7D32" if vs_meta >= 0 else "#CC0000"
            meta_mes_str  = fmt(meta_mes_rede)
            meta_hoje_str = fmt(meta_ate_hoje)
        else:
            meta_ate_hoje = 0
            vs_meta_str   = "—"
            vs_meta_cor   = "var(--subtexto)"
            meta_mes_str  = "—"
            meta_hoje_str = "—"

        # Mini chart (últimos 12 meses ou todos se < 12)
        _mini_raw   = labels_todos[-12:]
        mini_keys   = [f"{l[3:]}-{l[:2]}" for l in _mini_raw]
        mini_labels = [f"{MESES_PT_ABR[int(k[5:])-1]}/{k[:4]}" for k in mini_keys]
        mini_rede   = [totais_mes.get(k, 0) for k in mini_keys]
        mini_meta   = [meta_rede_por_mes.get(k) for k in mini_keys]

        # mês anterior para tabela
        partes_atual = chave_atual.split("-") if chave_atual else ["2026","01"]
        ano_a, mes_a = int(partes_atual[0]), int(partes_atual[1])
        chave_ant = f"{ano_a}-{mes_a-1:02d}" if mes_a > 1 else f"{ano_a-1}-12"

        # JSON para JS
        leads_unidades_list = sorted(unidades.keys())

        leads_data_js = "{\n"
        leads_data_js += f'  "rede": {_json.dumps(serie_rede())},\n'
        leads_data_js += '  "regioes": {\n'
        for reg in ["Regiao 1","Regiao 2","Regiao 3","Sao Paulo"]:
            leads_data_js += f'    "{reg}": {_json.dumps(serie_regiao(reg))},\n'
        leads_data_js += '  },\n'
        leads_data_js += '  "unidades": {\n'
        for nome_u in leads_unidades_list:
            leads_data_js += f'    "{nome_u}": {_json.dumps(serie_unidade(nome_u))},\n'
        leads_data_js += '  }\n}'

        leads_metas_js = "{\n"
        leads_metas_js += f'  "rede": {_json.dumps(meta_serie_rede())},\n'
        leads_metas_js += '  "regioes": {\n'
        for reg in ["Regiao 1","Regiao 2","Regiao 3","Sao Paulo"]:
            leads_metas_js += f'    "{reg}": {_json.dumps(meta_serie_regiao(reg))},\n'
        leads_metas_js += '  },\n'
        leads_metas_js += '  "unidades": {\n'
        for nome_u in leads_unidades_list:
            leads_metas_js += f'    "{nome_u}": {_json.dumps(meta_serie_unidade(nome_u))},\n'
        leads_metas_js += '  }\n}'

        mini_chart_js = f"""
const miniLabels = {_json.dumps(mini_labels)};
const miniRede   = {_json.dumps(mini_rede)};
const miniMeta   = {_json.dumps(mini_meta)};
"""
        # Conversão (ICL) por unidade/mês
        conv_js = "{\n"
        for nome_u in leads_unidades_list:
            conv_u = unidades.get(nome_u, {}).get("conversao", {})
            conv_js += f'  "{nome_u}": {_json.dumps(conv_u)},\n'
        conv_js += "}"

        # Ratio dias úteis para meta proporcional (mês atual)
        du_mes      = dias_uteis_no_mes(hoje.year, hoje.month)
        du_passados = dias_uteis_ate(hoje.year, hoje.month, hoje.day)
        du_ratio    = round(du_passados / du_mes, 6) if du_mes else 1.0

        # Anos disponíveis para o filtro
        anos_disponiveis = sorted({l[3:] for l in labels_todos}, reverse=True)

        leads_chart_js = f"""
const leadsAllLabels  = {_json.dumps(labels_todos)};
const leadsAllData    = {leads_data_js};
const leadsAllMetas   = {leads_metas_js};
const leadsUnidades   = {_json.dumps(leads_unidades_list)};
const leadsConversao  = {conv_js};
const leadsCurrentKey = {_json.dumps(chave_atual)};
const leadsMetaRatio  = {du_ratio};
const leadsOrdem      = {_json.dumps(leads_unidades_list)};
"""
    else:
        chave_atual = ""
        chave_ant   = ""
        meta_mes_str  = "—"
        meta_hoje_str = "—"
        vs_meta_str   = "—"
        vs_meta_cor   = "var(--subtexto)"
        leads_mes_atual = 0
        gerado_em = "—"
        mes_atual_key = ""
        mini_chart_js = ""
        leads_chart_js = ""
        leads_unidades_list = []

    # ── Meta seguidores / interações ───────────────────────────────
    metas_int = metas_data.get("interacoes", {})

    # Usa o mês/dia da coleta do Instagram (data_fim), não hoje
    try:
        _fim_parts = ig[0]["data_fim"].split("/") if ig else None
        _data_fim  = date(int(_fim_parts[2]), int(_fim_parts[1]), int(_fim_parts[0])) if _fim_parts else hoje
    except Exception:
        _data_fim = hoje
    import calendar as _cal
    chave_mes_ig = _data_fim.strftime("%Y-%m")          # chave do mês dos dados IG
    _dias_mes    = _cal.monthrange(_data_fim.year, _data_fim.month)[1]
    _ig_ratio    = round(_data_fim.day / _dias_mes, 6)  # proporção: dia_fim / total_dias

    def _meta_ig_ate_hoje(meta_total):
        return round(meta_total * _ig_ratio) if meta_total else 0

    meta_seg_total = sum(metas_seg.get(p, {}).get(chave_mes_ig, 0) for p in PERFIS_ORDEM)
    meta_int_total = sum(metas_int.get(p, {}).get(chave_mes_ig, 0) for p in PERFIS_ORDEM)

    # Mês anterior para buscar fechamento
    _ano_ig, _mes_ig_n = _data_fim.year, _data_fim.month
    _mes_ant_n = _mes_ig_n - 1 if _mes_ig_n > 1 else 12
    _ano_ant   = _ano_ig if _mes_ig_n > 1 else _ano_ig - 1
    _mes_ant_key = f"{_mes_ant_n:02d}/{_ano_ant}"   # ex: "06/2026"
    fechamento_ant = seg_total_hist.get(_mes_ant_key, {})

    # Dias úteis do mês atual e passados até hoje
    _du_mes_ig   = dias_uteis_no_mes(_ano_ig, _mes_ig_n)
    _du_passados = dias_uteis_ate(_ano_ig, _mes_ig_n, hoje.day)

    def _meta_seg_perfil_hoje(nome):
        """Fechamento anterior + (diferença / du_mes) × du_passados"""
        meta_mes = metas_seg.get(nome, {}).get(chave_mes_ig)
        fech_ant = fechamento_ant.get(nome)
        if meta_mes is None or fech_ant is None or not _du_mes_ig:
            return meta_mes
        diferenca   = meta_mes - fech_ant
        meta_diaria = diferenca / _du_mes_ig
        return round(fech_ant + meta_diaria * _du_passados)

    # Meta até hoje total = soma das metas individuais por perfil
    meta_seg_ate_hoje = sum(
        (_meta_seg_perfil_hoje(p) or 0) for p in PERFIS_ORDEM
        if metas_seg.get(p, {}).get(chave_mes_ig)
    )
    # Interações acumulam no período → prorrateio simples por dia
    meta_int_ate_hoje = _meta_ig_ate_hoje(meta_int_total)

    # ── ABA VISÃO GERAL ────────────────────────────────────────────
    if leads:
        leads_cards = f"""
      <div class="section-label">Leads CRM Pacto &nbsp;·&nbsp; atualizado em {gerado_em}</div>
      <div style="display:grid; grid-template-columns:2fr 1fr; gap:10px; margin-bottom:8px">
        <div class="big-card blue" style="padding:14px 18px">
          <div class="bc-label">Tendência — Total Rede</div>
          <canvas id="chart-leads-mini" style="max-height:140px"></canvas>
        </div>
        <div class="big-card blue">
          <div class="bc-label">Leads {mes_atual_key}</div>
          <div class="bc-valor">{fmt(leads_mes_atual)}</div>
          <div class="bc-sub">mês atual · parcial</div>
          <div class="bc-detail">
            <div class="bc-detail-row"><span>Meta até hoje</span><span>{meta_hoje_str}</span></div>
            <div class="bc-detail-row"><span>vs meta</span><span style="color:{vs_meta_cor};font-weight:700">{vs_meta_str}</span></div>
            <div class="bc-detail-row"><span>Meta do mês</span><span>{meta_mes_str}</span></div>
          </div>
        </div>
      </div>
      <div style="text-align:right;margin-bottom:16px">
        <button class="ver-mais-btn ver-mais-blue" onclick="showTab('leads','tab-leads')">Ver mais →</button>
      </div>"""
    else:
        leads_cards = ""

    # ── Construção das linhas de sub-meta por perfil ───────────────
    n_perfis = len(ig)
    perfis_label = f"soma dos {n_perfis} perfis"

    def _sub_meta_row(nome, valor, meta_dict, proratar=False):
        meta_p      = meta_dict.get(nome, {}).get(chave_mes_ig, None)
        meta_p_hoje = round(meta_p * _ig_ratio) if (meta_p and proratar) else meta_p
        if meta_p_hoje:
            diff = valor - meta_p_hoje
            cor  = "#2E7D32" if diff >= 0 else "#CC0000"
            sinal = "+" if diff >= 0 else ""
            detalhe = (
                f'<span style="color:var(--subtexto);font-size:11px"> meta: {fmt(meta_p_hoje)}</span>'
                f'<span style="color:{cor};font-size:11px;font-weight:600"> {sinal}{fmt(diff)}</span>'
            )
        else:
            detalhe = ""
        return f'<div class="bc-detail-row"><span>{nome}</span><span>{fmt(valor)}{detalhe}</span></div>'

    def _sub_meta_rows_seg():
        rows = []
        for p in ig:
            nome  = p["nome"]
            valor = p["total"] or 0
            meta_hoje = _meta_seg_perfil_hoje(nome)
            if meta_hoje is not None:
                diff = valor - meta_hoje
                cor  = "#2E7D32" if diff >= 0 else "#CC0000"
                sinal = "+" if diff >= 0 else ""
                detalhe = (
                    f'<span style="color:var(--subtexto);font-size:11px"> meta: {fmt(meta_hoje)}</span>'
                    f'<span style="color:{cor};font-size:11px;font-weight:600"> {sinal}{fmt(diff)}</span>'
                )
            else:
                detalhe = ""
            rows.append(f'<div class="bc-detail-row"><span>{nome}</span><span>{fmt(valor)}{detalhe}</span></div>')
        return "".join(rows)

    def _sub_meta_rows_int():
        return "".join(_sub_meta_row(p["nome"], p["total_int"] or 0, metas_int, proratar=True) for p in ig)

    # bloco meta visível ao lado do número grande
    def _meta_badge(valor_atual, meta_ate_hoje, meta_mes):
        if not meta_ate_hoje:
            return ""
        diff = valor_atual - meta_ate_hoje
        cor  = "#2E7D32" if diff >= 0 else "#CC0000"
        sinal = "+" if diff >= 0 else ""
        return (
            f'<div class="bc-valor">{fmt(valor_atual)}</div>'
            f'<div class="bc-sub">{perfis_label} · meta do mês: {fmt(meta_mes)}</div>'
            f'<div style="margin-top:6px;padding:6px 10px;border-radius:8px;background:var(--cinza2);font-size:12px;color:var(--subtexto)">'
            f'Meta até dia {_data_fim.day}: <strong style="color:var(--texto)">{fmt(meta_ate_hoje)}</strong>'
            f'&nbsp;&nbsp;<span style="color:{cor};font-weight:700">{sinal}{fmt(diff)}</span>'
            f'</div>'
        )

    # Seguidores: meta até hoje calculada pela fórmula dias úteis
    if meta_seg_ate_hoje > 0:
        diff_seg  = total_seguidores - meta_seg_ate_hoje
        cor_seg   = "#2E7D32" if diff_seg >= 0 else "#CC0000"
        sinal_seg = "+" if diff_seg >= 0 else ""
        seg_header_html = (
            f'<div class="bc-valor">{fmt(total_seguidores)}</div>'
            f'<div class="bc-sub">{perfis_label} · meta do mês: {fmt(meta_seg_total)}</div>'
            f'<div style="margin-top:6px;padding:6px 10px;border-radius:8px;background:var(--cinza2);font-size:12px;color:var(--subtexto)">'
            f'Meta até hoje ({_du_passados} d.u.): <strong style="color:var(--texto)">{fmt(meta_seg_ate_hoje)}</strong>'
            f'&nbsp;&nbsp;<span style="color:{cor_seg};font-weight:700">{sinal_seg}{fmt(diff_seg)}</span>'
            f'</div>'
        )
    else:
        seg_header_html = f'<div class="bc-valor">{fmt(total_seguidores)}</div><div class="bc-sub">{perfis_label}</div>'

    # Interações: acumulam no mês → prorateio faz sentido
    int_header_html = _meta_badge(total_interacoes, meta_int_ate_hoje, meta_int_total) or (
        f'<div class="bc-valor">{fmt(total_interacoes)}</div><div class="bc-sub">{perfis_label}</div>'
    )

    aba_geral = f"""
    <div class="cards-section">
      {leads_cards}

      <div class="section-label" style="margin-top:12px">Instagram &nbsp;·&nbsp; {semana_ig} &nbsp;·&nbsp; {periodo_ig}</div>
      <div class="big-cards">
        <div class="big-card purple">
          <div class="bc-label">Total Seguidores</div>
          {seg_header_html}
          <div class="bc-detail">
            {_sub_meta_rows_seg()}
          </div>
        </div>
        <div class="big-card purple">
          <div class="bc-label">Total Interações</div>
          {int_header_html}
          <div class="bc-detail">{_sub_meta_rows_int()}</div>
        </div>
      </div>
      <div style="text-align:right;margin-bottom:16px">
        <button class="ver-mais-btn ver-mais-purple" onclick="showTab('instagram','tab-instagram')">Ver mais →</button>
      </div>

      <div class="section-label" style="margin-top:12px">Clube de Vantagens &nbsp;·&nbsp; atualizado em {resumo['data']}</div>
      <div class="big-cards">
        <div class="big-card green">
          <div class="bc-label">Parceiros</div>
          <div class="bc-valor">{fmt(resumo['total_parceiros'])}</div>
          <div class="bc-sub">{resumo['parc_com_usuario']} com usuário · {resumo['parc_sem_usuario']} sem</div>
        </div>
        <div class="big-card green">
          <div class="bc-label">Promoções</div>
          <div class="bc-valor">{fmt(resumo['total_promocoes'])}</div>
          <div class="bc-sub">cadastradas</div>
        </div>
        <div class="big-card green">
          <div class="bc-label">Cupons Gerados</div>
          <div class="bc-valor">{fmt(resumo['cupons_gerados'])}</div>
          <div class="bc-sub">desde o início</div>
        </div>
        <div class="big-card orange">
          <div class="bc-label">Cupons Utilizados</div>
          <div class="bc-valor">{fmt(resumo['cupons_utilizados'])}</div>
          <div class="bc-sub">resgates confirmados</div>
        </div>
      </div>
      <div style="text-align:right">
        <button class="ver-mais-btn ver-mais-green" onclick="showTab('parceiros','tab-parceiros')">Ver mais →</button>
      </div>
    </div>"""

    # ── ABA INSTAGRAM ──────────────────────────────────────────────
    todos_indicadores = {
        "total_seguidores":  "Total Seguidores",
        "crescimento":       "Crescimento Líquido",
        "total_int":         "Total Interações (API)",
        "curtidas":          "Curtidas",
        "comentarios":       "Comentários",
        "compartilhamentos": "Compartilhamentos",
        "salvamentos":       "Salvamentos",
        "posts":             "Posts no Período",
    }

    def tem_dados(ind):
        return any(v is not None for p in PERFIS_ORDEM for v in ig_hist[ind][p].values())

    indicadores_labels = {k: v for k, v in todos_indicadores.items() if tem_dados(k)}

    chart_labels_ig = _json.dumps(ig_meses)

    chart_datasets_ig = {}
    for ind in indicadores_labels:
        datasets = []
        for perfil in PERFIS_ORDEM:
            valores = [ig_hist[ind][perfil].get(m) for m in ig_meses]
            datasets.append({
                "label": perfil, "data": valores,
                "borderColor": CORES_PERFIS[perfil],
                "backgroundColor": CORES_PERFIS[perfil] + "22",
                "borderWidth": 2, "pointRadius": 5, "pointHoverRadius": 7,
                "tension": 0.3, "fill": False, "spanGaps": True,
            })
        chart_datasets_ig[ind] = _json.dumps(datasets)

    chart_datasets_js = "{\n" + ",\n".join(
        f'  "{k}": {v}' for k, v in chart_datasets_ig.items()
    ) + "\n}"

    opcoes_select = "\n".join(
        f'<option value="{k}">{v}</option>' for k, v in indicadores_labels.items()
    )

    ig_rows = ""
    for p in ig:
        nome  = p["nome"]
        sinal = "+" if isinstance(p["crescimento"], (int, float)) and p["crescimento"] > 0 else ""
        meta_p    = metas_seg.get(nome, {}).get(chave_mes_ig)
        meta_hoje = _meta_seg_perfil_hoje(nome)
        if meta_p and p["total"] is not None:
            diff_p  = p["total"] - (meta_hoje or meta_p)
            cor_p   = "#2E7D32" if diff_p >= 0 else "#CC0000"
            sinal_p = "+" if diff_p >= 0 else ""
            meta_cell = (
                f'<span style="color:var(--subtexto)">{fmt(meta_hoje or meta_p)}</span>'
                f'&nbsp;<span style="color:{cor_p};font-weight:700;font-size:11px">{sinal_p}{fmt(diff_p)}</span>'
            )
        else:
            meta_cell = "—"
        ig_rows += f"""
        <tr>
          <td class="td-nome">{nome}</td>
          <td class="num">{fmt(p['total'])}</td>
          <td class="num">{meta_cell}</td>
          <td class="num">{sinal}{fmt(p['crescimento'])}</td>
          <td class="num blue-txt">{fmt(p['total_int'])}</td>
          <td class="num">{fmt(p['curtidas'])}</td>
          <td class="num">{fmt(p['comentarios'])}</td>
          <td class="num">{fmt(p['compartilhamentos'])}</td>
          <td class="num">{fmt(p['salvamentos'])}</td>
          <td class="num">{fmt(p['posts'])}</td>
        </tr>"""

    aba_instagram = f"""
    <div class="back-bar">
      <button class="back-btn" onclick="showTab('geral','tab-geral')">← Visão Geral</button>
      <span class="back-sep">/</span><span class="back-label">Instagram</span>
    </div>
    <div class="period-bar">
      <span class="period-badge">{semana_ig}</span>
      {periodo_ig}
    </div>
    <div class="panel">
      <div class="panel-header">
        <div class="dot blue-dot"></div>
        <div><h2>Métricas por Perfil</h2><p>Acumulado do mês até a última coleta</p></div>
      </div>
      <div class="table-wrap">
        <table id="tbl-ig">
          <thead>
            <tr>
              <th class="th-sort" onclick="sortTable('tbl-ig',0)">Perfil</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',1)">Seguidores</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',2)">Meta / vs</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',3)">Crescimento</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',4)">Total Interações <span class="info-badge" title="Inclui stories, reels e ações de perfil">i</span></th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',5)">Curtidas</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',6)">Comentários</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',7)">Compartilh.</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',8)">Salvamentos</th>
              <th class="th-sort num" onclick="sortTable('tbl-ig',9)">Posts</th>
            </tr>
          </thead>
          <tbody>{ig_rows}</tbody>
        </table>
      </div>
      <div style="padding:8px 16px 10px;font-size:11px;color:var(--subtexto)">
        ⓘ Total Interações vem diretamente da API da Meta e inclui interações além das colunas acima: respostas a stories, tapadas (forward/back), plays de reels, visitas ao perfil e cliques em links. Os dados detalhados (curtidas, comentários, etc.) são coletados por publicação individual.
      </div>
    </div>
    <div class="panel" style="margin-bottom:24px">
      <div class="panel-header">
        <div class="dot blue-dot"></div>
        <div><h2>Evolução por Indicador</h2><p>Acumulado mensal — último registro de cada mês</p></div>
        <div style="margin-left:auto;display:flex;gap:8px;align-items:center;flex-wrap:wrap">
          <div class="view-btns">
            <button class="pill active" id="ig-btn-perfil" onclick="setIgView('perfil',this)">Por Perfil</button>
            <button class="pill" id="ig-btn-rede" onclick="setIgView('rede',this)">Visão Rede</button>
          </div>
          <select id="ig-select" onchange="updateIgChart()" style="font-size:13px;padding:6px 12px;border-radius:8px;border:1px solid var(--borda);background:var(--cinza2);color:var(--texto);cursor:pointer">
            {opcoes_select}
          </select>
        </div>
      </div>
      <div style="padding:24px">
        <canvas id="ig-chart" height="90"></canvas>
      </div>
    </div>"""

    # ── ABA PARCEIROS ──────────────────────────────────────────────
    linhas_ranking = ""
    for i, p in enumerate(ranking, 1):
        linhas_ranking += f"""
        <tr>
          <td class="rank">#{i}</td>
          <td class="td-nome">{p['nome']}</td>
          <td class="num">{p['gerados']}</td>
          <td class="num green-txt">{p['utilizados']}</td>
          <td class="num"><span class="badge">{p['taxa']}%</span></td>
        </tr>"""

    linhas_mensal = ""
    for mk, vals in hist_mensal_sorted:
        taxa = round(vals["utilizados"] / vals["gerados"] * 100, 1) if vals["gerados"] else 0
        # data-sort como YYYYMM para ordenação cronológica correta
        try:
            _m, _y = mk.split("/")
            sort_mes = f"{_y}{_m}"
        except Exception:
            sort_mes = mk
        linhas_mensal += f"""
        <tr>
          <td class="td-nome" data-sort="{sort_mes}">{fmt_mes(mk)}</td>
          <td class="num">{vals['gerados']}</td>
          <td class="num green-txt">{vals['utilizados']}</td>
          <td class="num"><span class="badge">{taxa}%</span></td>
        </tr>"""

    aba_parceiros = f"""
    <div class="back-bar">
      <button class="back-btn" onclick="showTab('geral','tab-geral')">← Visão Geral</button>
      <span class="back-sep">/</span><span class="back-label">Parceiros</span>
    </div>
    <div class="period-bar">Dados extraídos de parceirogreenlife.com.br &nbsp;·&nbsp; atualizado em {resumo['data']}</div>
    <div class="big-cards" style="margin-bottom:24px">
      <div class="big-card green">
        <div class="bc-label">Parceiros</div>
        <div class="bc-valor">{fmt(resumo['total_parceiros'])}</div>
        <div class="bc-sub">{resumo['parc_com_usuario']} com usuário · {resumo['parc_sem_usuario']} sem</div>
      </div>
      <div class="big-card green">
        <div class="bc-label">Promoções</div>
        <div class="bc-valor">{fmt(resumo['total_promocoes'])}</div>
        <div class="bc-sub">cadastradas</div>
      </div>
      <div class="big-card green">
        <div class="bc-label">Cupons Gerados</div>
        <div class="bc-valor">{fmt(resumo['cupons_gerados'])}</div>
        <div class="bc-sub">desde o início</div>
      </div>
      <div class="big-card orange">
        <div class="bc-label">Cupons Utilizados</div>
        <div class="bc-valor">{fmt(resumo['cupons_utilizados'])}</div>
        <div class="bc-sub">resgates confirmados</div>
      </div>
    </div>
    <div class="two-col">
      <div class="panel">
        <div class="panel-header">
          <div class="dot"></div>
          <div><h2>Cupons por Parceiro</h2><p>Parceiros com pelo menos 1 resgate</p></div>
        </div>
        {"<div class='table-wrap'><table id='tbl-parceiros'><thead><tr><th></th><th class='th-sort' onclick=\"sortTable('tbl-parceiros',1)\">Parceiro</th><th class='th-sort num' onclick=\"sortTable('tbl-parceiros',2)\">Gerados</th><th class='th-sort num' onclick=\"sortTable('tbl-parceiros',3)\">Utilizados</th><th class='th-sort num' onclick=\"sortTable('tbl-parceiros',4)\">Taxa</th></tr></thead><tbody>" + linhas_ranking + "</tbody></table></div>" if ranking else "<div class='empty'>Nenhum cupom utilizado registrado.</div>"}
        <div style="padding:8px 16px 12px;font-size:11px;color:var(--subtexto)">Ranking acumulado — sem quebra mensal por parceiro</div>
      </div>
      <div class="panel">
        <div class="panel-header">
          <div class="dot"></div>
          <div><h2>Cupons por Mês</h2><p>Cupons gerados e utilizados por mês</p></div>
        </div>
        <div class="table-wrap">
          <table id="tbl-mensal">
            <thead>
              <tr>
                <th class="th-sort" onclick="sortTable('tbl-mensal',0)">Mês</th>
                <th class="th-sort num" onclick="sortTable('tbl-mensal',1)">Gerados</th>
                <th class="th-sort num" onclick="sortTable('tbl-mensal',2)">Utilizados</th>
                <th class="th-sort num" onclick="sortTable('tbl-mensal',3)">Taxa</th>
              </tr>
            </thead>
            <tbody>{linhas_mensal}</tbody>
          </table>
        </div>
      </div>
    </div>"""

    # ── ABA LEADS ─────────────────────────────────────────────────
    if leads:
        unidades_mes = sorted([
            (nome, d["regiao"], d["meses"].get(chave_atual, 0), d["meses"].get(chave_ant, 0))
            for nome, d in unidades.items()
        ], key=lambda x: -x[2])

        linhas_leads = ""
        for i, (nome, regiao, atual_v, ant_v) in enumerate(unidades_mes, 1):
            if ant_v > 0:
                var_pct = round((atual_v - ant_v) / ant_v * 100)
                cor_var = "#CC0000" if var_pct < 0 else "#2E7D32"
                var_str = f'<td class="num" style="color:{cor_var}">{("+" if var_pct>0 else "")}{var_pct}%</td>'
            else:
                var_str = '<td class="num" style="color:var(--subtexto)">N/A</td>'
            linhas_leads += f'<tr><td class="rank">#{i}</td><td class="td-nome">{nome}</td><td>{regiao.replace("Regiao","Região")}</td><td class="num blue-txt">{fmt(atual_v)}</td><td class="num">{fmt(ant_v)}</td>{var_str}</tr>\n'

        total_mes_atual = totais_mes.get(chave_atual, 0)
        total_mes_ant   = totais_mes.get(chave_ant, 0)

        unit_options = "\n".join(
            f'<option value="{u}">{u}</option>' for u in leads_unidades_list
        )

        anos_options = "\n".join(f'<option value="{a}">{a}</option>' for a in anos_disponiveis)
        meses_options = "\n".join(
            f'<option value="{m:02d}">{MESES_PT_ABR[m-1]}</option>' for m in range(1,13)
        )

        aba_leads = f"""
    <div class="back-bar">
      <button class="back-btn" onclick="showTab('geral','tab-geral')">← Visão Geral</button>
      <span class="back-sep">/</span><span class="back-label">Leads Pacto</span>
    </div>
    <div class="period-bar">
      <span class="period-badge">Jan → {mes_atual_key}</span>
      CRM · todos os canais · 24 unidades · atualizado em {gerado_em}
    </div>
    <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:18px;padding:14px 18px;background:var(--cinza2);border-radius:12px;border:1px solid var(--borda)">
      <div class="view-btns" style="margin-right:4px">
        <button class="pill lv-btn active" id="lv-rede"    onclick="setLeadsView('rede',this)">Rede</button>
        <button class="pill lv-btn"        id="lv-regiao"  onclick="setLeadsView('regiao',this)">Por Região</button>
        <button class="pill lv-btn"        id="lv-unidade" onclick="setLeadsView('unidade',this)">Por Unidade</button>
      </div>
      <select id="leads-reg-sel" style="display:none;font-size:13px;padding:6px 12px;border-radius:8px;border:1px solid var(--borda);background:var(--card);color:var(--texto);cursor:pointer" onchange="leadsReg=this.value;onLeadsFilter()">
        <option value="Regiao 1">Região 1</option>
        <option value="Regiao 2">Região 2</option>
        <option value="Regiao 3">Região 3</option>
        <option value="Sao Paulo">São Paulo</option>
      </select>
      <select id="leads-unit-sel" style="display:none;font-size:13px;padding:6px 12px;border-radius:8px;border:1px solid var(--borda);background:var(--card);color:var(--texto);cursor:pointer" onchange="leadsUnit=this.value;onLeadsFilter()">
        {unit_options}
      </select>
      <div style="margin-left:auto;display:flex;gap:8px;align-items:center">
        <select id="lv-year" style="font-size:13px;padding:6px 12px;border-radius:8px;border:1px solid var(--borda);background:var(--card);color:var(--texto);cursor:pointer" onchange="onLeadsFilter()">
          {anos_options}
        </select>
        <select id="lv-month" style="font-size:13px;padding:6px 12px;border-radius:8px;border:1px solid var(--borda);background:var(--card);color:var(--texto);cursor:pointer" onchange="onLeadsFilter()">
          {meses_options}
        </select>
      </div>
    </div>
    <div class="panel" style="margin-bottom:20px">
      <div class="panel-header">
        <div class="dot blue-dot"></div>
        <div><h2>Evolução de Leads</h2><p>Série histórica completa com meta</p></div>
      </div>
      <div style="padding:20px"><canvas id="chart-leads-regiao" height="110"></canvas></div>
    </div>
    <div class="panel" style="margin-bottom:20px">
      <div class="panel-header" style="padding-bottom:0">
        <div class="dot blue-dot"></div>
        <div><h2>Leads por Unidade</h2><p id="lv-period-label">Filtrado por mês</p></div>
      </div>
      <div style="padding:16px 20px 8px"><canvas id="chart-leads-units"></canvas></div>
    </div>
    <div class="panel">
      <div class="panel-header">
        <div class="dot blue-dot"></div>
        <div><h2>Resultado por Unidade</h2><p id="tbl-leads-subtitle">Mês selecionado</p></div>
      </div>
      <div class="table-wrap">
        <table id="tbl-leads">
          <thead>
            <tr>
              <th style="width:32px"></th>
              <th class="th-sort" onclick="sortTable('tbl-leads',1)">Unidade</th>
              <th class="th-sort" onclick="sortTable('tbl-leads',2)">Região</th>
              <th class="th-sort num" onclick="sortTable('tbl-leads',3)">Leads</th>
              <th class="th-sort num" onclick="sortTable('tbl-leads',4)">Meta/dia</th>
              <th class="th-sort num" onclick="sortTable('tbl-leads',5)">Falta</th>
              <th class="th-sort num" onclick="sortTable('tbl-leads',6)">Conv.</th>
              <th class="th-sort num" onclick="sortTable('tbl-leads',7)">ICL %</th>
            </tr>
          </thead>
          <tbody id="tbl-leads-body"></tbody>
          <tfoot id="tbl-leads-foot"></tfoot>
        </table>
      </div>
    </div>"""
    else:
        aba_leads = "<div class='empty'>Dados de leads não encontrados. Execute bi_crm_leads.py.</div>"

    # ── JAVASCRIPT ────────────────────────────────────────────────
    chart_script = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<script>
Chart.register(ChartDataLabels);

// ── Sort tables ────────────────────────────────────────────────
function parseSortVal(s) {{
  const t = s.trim().replace(/[^0-9\\-,\\.]/g,'');
  if (!t || t==='-') return -Infinity;
  return parseFloat(t.replace(/\\.(?=\\d{{3}})/g,'').replace(',','.')) || 0;
}}

function sortTable(tblId, col) {{
  const tbl = document.getElementById(tblId);
  if (!tbl) return;
  const ths   = tbl.querySelectorAll('thead th');
  const tbody = tbl.querySelector('tbody');
  const asc   = ths[col].dataset.sortDir !== 'asc';
  ths.forEach(t => {{ delete t.dataset.sortDir; t.classList.remove('s-asc','s-desc'); }});
  ths[col].dataset.sortDir = asc ? 'asc' : 'desc';
  ths[col].classList.add(asc ? 's-asc' : 's-desc');
  const rows = Array.from(tbody.rows);
  rows.sort((a,b) => {{
    const cellA = a.cells[col], cellB = b.cells[col];
    const av = cellA?.dataset.sort || cellA?.textContent || '';
    const bv = cellB?.dataset.sort || cellB?.textContent || '';
    const an = parseSortVal(av), bn = parseSortVal(bv);
    if (isFinite(an) && isFinite(bn)) return asc ? an-bn : bn-an;
    return asc ? av.localeCompare(bv,'pt-BR') : bv.localeCompare(av,'pt-BR');
  }});
  rows.forEach(r => tbody.appendChild(r));
}}

// ── Mini chart (Visão Geral) ───────────────────────────────────
{mini_chart_js}
const miniColors = miniRede.map((v,i) => {{
  const m = miniMeta[i];
  if (v==null||m==null) return 'rgba(158,158,158,0.45)';
  return v>=m ? 'rgba(46,125,50,0.75)' : 'rgba(198,40,40,0.75)';
}});
const miniBorders = miniRede.map((v,i) => {{
  const m = miniMeta[i];
  if (v==null||m==null) return '#9E9E9E';
  return v>=m ? '#2E7D32' : '#C62828';
}});

let miniChart = null;
function initMiniChart() {{
  const ctx = document.getElementById('chart-leads-mini');
  if (!ctx || miniChart) return;

  const _allVals = [...miniRede.filter(v=>v!=null), ...miniMeta.filter(v=>v!=null)];
  const _min = _allVals.length ? Math.max(0, Math.floor(Math.min(..._allVals) * 0.90)) : 0;

  miniChart = new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels: miniLabels,
      datasets: [
        {{ label: 'Leads', data: miniRede, type: 'bar',
           backgroundColor: miniColors, borderColor: miniBorders,
           borderWidth: 1.5, borderRadius: 4, borderSkipped: false, order: 1 }},
        {{ label: 'Meta', data: miniMeta, type: 'line',
           borderColor: '#9E9E9E', borderDash: [5,4], backgroundColor: 'transparent',
           borderWidth: 1.5, pointRadius: 0, tension: 0, fill: false, spanGaps: true, order: 0 }},
      ]
    }},
    options: {{
      responsive: true,
      plugins: {{
        legend: {{ display: false }},
        datalabels: {{
          display: ctx => ctx.datasetIndex === 0 && ctx.dataset.data[ctx.dataIndex] != null,
          align: 'top',
          anchor: 'end',
          font: {{ size: 10, weight: '700' }},
          color: ctx => miniBorders[ctx.dataIndex],
          formatter: v => v ? v.toLocaleString('pt-BR') : '',
          padding: {{ bottom: 2 }},
        }},
        tooltip: {{
          callbacks: {{
            title: ctx => miniLabels[ctx[0].dataIndex],
            label: ctx => {{
              const v = ctx.raw;
              if (ctx.datasetIndex === 1) return 'Meta: ' + (v||0).toLocaleString('pt-BR');
              const m = miniMeta[ctx.dataIndex];
              const diff = (m!=null && v!=null) ? v - m : null;
              const lines = ['Leads: '+(v||0).toLocaleString('pt-BR')];
              if (m!=null) lines.push('Meta: '+m.toLocaleString('pt-BR'));
              if (diff!=null) lines.push(diff>=0
                ? '✓ +'+ diff.toLocaleString('pt-BR')+' acima da meta'
                : '✗ ' + diff.toLocaleString('pt-BR')+' abaixo da meta');
              return lines;
            }},
          }}
        }},
      }},
      scales: {{
        x: {{
          display: true,
          ticks: {{ font: {{ size: 10 }}, color: '#9CA3AF', maxRotation: 0 }},
          grid: {{ display: false }},
        }},
        y: {{
          display: true,
          beginAtZero: false,
          suggestedMin: _min,
          ticks: {{
            maxTicksLimit: 5,
            font: {{ size: 10 }},
            color: '#9CA3AF',
            callback: v => v.toLocaleString('pt-BR'),
          }},
          grid: {{ color: '#F0F2F5' }},
        }},
      }},
      layout: {{ padding: {{ top: 20 }} }},
    }},
  }});
}}

// ── Instagram chart ────────────────────────────────────────────
const igLabels   = {chart_labels_ig};
const igDatasets = {chart_datasets_js};
const igNames    = {_json.dumps(indicadores_labels)};
let igChart     = null;
let igViewMode  = 'perfil';

function setIgView(view, btn) {{
  igViewMode = view;
  document.querySelectorAll('#ig-btn-perfil,#ig-btn-rede').forEach(b => b.classList.remove('active'));
  if (btn) btn.classList.add('active');
  updateIgChart();
}}

function getIgDatasets(ind) {{
  const base = igDatasets[ind] || [];
  if (igViewMode === 'perfil') return base;
  const agg = igLabels.map((_,i) => {{
    const vals = base.map(d => d.data[i]).filter(v => v != null);
    return vals.length ? vals.reduce((a,b) => a+b, 0) : null;
  }});
  return [{{ label: 'Total Rede', data: agg,
    borderColor: '#1B3A2F', backgroundColor: '#1B3A2F20',
    borderWidth: 3, pointRadius: 5, pointHoverRadius: 7, tension: 0.3, fill: false, spanGaps: true }}];
}}

function updateIgChart() {{
  const ind = document.getElementById('ig-select').value;
  const datasets = getIgDatasets(ind);
  if (igChart) {{
    igChart.data.labels = igLabels;
    igChart.data.datasets = datasets;
    igChart.options.plugins.title.text = igNames[ind];
    igChart.update(); return;
  }}
  igChart = new Chart(document.getElementById('ig-chart'), {{
    type: 'line',
    data: {{ labels: igLabels, datasets: datasets }},
    options: {{
      responsive: true,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{
        title: {{ display: true, text: igNames[ind], font: {{ size: 14, weight: '600' }}, padding: {{ bottom: 16 }} }},
        legend: {{ position: 'bottom', labels: {{ usePointStyle: true, padding: 20, font: {{ size: 13 }} }} }},
        datalabels: {{
          display: ctx => ctx.dataset.data[ctx.dataIndex] != null,
          align: 'top', anchor: 'end', offset: 4,
          font: {{ size: 10, weight: '700' }},
          color: '#fff',
          backgroundColor: ctx => ctx.dataset.borderColor,
          borderRadius: 4,
          padding: {{ top: 3, bottom: 3, left: 5, right: 5 }},
          formatter: v => v == null ? '' : v.toLocaleString('pt-BR'),
          clip: false,
        }},
      }},
      scales: {{
        x: {{ grid: {{ color: '#F0F2F5' }}, ticks: {{ font: {{ size: 12 }} }} }},
        y: {{ grid: {{ color: '#F0F2F5' }},
              ticks: {{ font: {{ size: 12 }}, callback: v => v.toLocaleString('pt-BR') }},
              beginAtZero: false }},
      }},
    }},
  }});
}}

// ── Leads chart ────────────────────────────────────────────────
{leads_chart_js}
let leadsChart = null;
let leadsView  = 'rede';
let leadsUnit  = (typeof leadsUnidades !== 'undefined' && leadsUnidades.length) ? leadsUnidades[0] : '';
let leadsReg   = 'Regiao 1';

function setLeadsView(view, btn) {{
  leadsView = view;
  document.querySelectorAll('.lv-btn').forEach(b => b.classList.remove('active'));
  if (btn) btn.classList.add('active');
  const selU = document.getElementById('leads-unit-sel');
  const selR = document.getElementById('leads-reg-sel');
  if (selU) selU.style.display = view === 'unidade' ? 'inline-block' : 'none';
  if (selR) selR.style.display = view === 'regiao'  ? 'inline-block' : 'none';
  onLeadsFilter();
}}

function renderLeadsChart() {{
  if (typeof leadsAllLabels === 'undefined') return;
  const COLORS = {{'Regiao 1':'#1565C0','Regiao 2':'#2E7D32','Regiao 3':'#F57C00','Sao Paulo':'#6A1B9A'}};
  const RNAMES = {{'Regiao 1':'Região 1','Regiao 2':'Região 2','Regiao 3':'Região 3','Sao Paulo':'São Paulo'}};
  const datasets = [];

  if (leadsView === 'rede') {{
    datasets.push({{ label: 'Total Rede', data: leadsAllData.rede,
      borderColor:'#1565C0', backgroundColor:'#1565C022',
      borderWidth:2, pointRadius:3, tension:0.3, fill:false, spanGaps:true }});
    const mRede = leadsAllMetas.rede;
    if (mRede && mRede.some(v => v!=null))
      datasets.push({{ label:'Meta', data:mRede,
        borderColor:'#9E9E9E', borderDash:[6,4], backgroundColor:'transparent',
        borderWidth:2, pointRadius:0, tension:0.3, fill:false, spanGaps:true,
        datalabels:{{display:false}} }});

  }} else if (leadsView === 'regiao') {{
    const r = leadsReg;
    const cor = COLORS[r] || '#1565C0';
    datasets.push({{ label: RNAMES[r]||r, data: leadsAllData.regioes[r]||[],
      borderColor:cor, backgroundColor:cor+'22',
      borderWidth:2, pointRadius:3, tension:0.3, fill:false, spanGaps:true }});
    const mR = leadsAllMetas.regioes[r];
    if (mR && mR.some(v=>v!=null))
      datasets.push({{ label:'Meta', data:mR,
        borderColor:'#9E9E9E', borderDash:[6,4], backgroundColor:'transparent',
        borderWidth:1.5, pointRadius:0, tension:0.3, fill:false, spanGaps:true,
        datalabels:{{display:false}} }});

  }} else if (leadsView === 'unidade') {{
    const u = leadsUnit;
    datasets.push({{ label:u, data:leadsAllData.unidades[u]||[],
      borderColor:'#1565C0', backgroundColor:'#1565C022',
      borderWidth:2, pointRadius:4, tension:0.3, fill:false, spanGaps:true }});
    const mU = leadsAllMetas.unidades[u];
    if (mU && mU.some(v=>v!=null))
      datasets.push({{ label:'Meta', data:mU,
        borderColor:'#9E9E9E', borderDash:[6,4], backgroundColor:'transparent',
        borderWidth:2, pointRadius:0, tension:0.3, fill:false, spanGaps:true,
        datalabels:{{display:false}} }});
  }}

  if (leadsChart) {{
    leadsChart.data.labels   = leadsAllLabels;
    leadsChart.data.datasets = datasets;
    leadsChart.update();
  }} else {{
    leadsChart = new Chart(document.getElementById('chart-leads-regiao'), {{
      type: 'line',
      data: {{ labels: leadsAllLabels, datasets: datasets }},
      options: {{
        responsive: true,
        interaction: {{ mode: 'index', intersect: false }},
        plugins: {{
          legend: {{
            position: 'top',
            labels: {{ usePointStyle:true, padding:16, font:{{ size:12 }} }},
            onClick: () => {{}},
          }},
          datalabels: {{
            align:'top', anchor:'end',
            font:{{ size:10, weight:'600' }},
            formatter: v => v ? v.toLocaleString('pt-BR') : '',
            color: ctx => ctx.dataset.borderColor,
          }},
        }},
        scales: {{
          x: {{ grid:{{ color:'#F0F2F5' }}, ticks:{{ font:{{ size:11 }} }} }},
          y: {{ beginAtZero:false, grid:{{ color:'#F0F2F5' }},
                ticks:{{ callback: v => v.toLocaleString('pt-BR') }} }},
        }},
        layout: {{ padding: {{ top: 20 }} }},
      }},
    }});
  }}
}}

function initLeadsChart() {{
  const ctx = document.getElementById('chart-leads-regiao');
  if (!ctx) return;
  renderLeadsChart();
}}

// ── Filtro Mês / Gráfico Unidades / Tabela ─────────────────────
let unitsChart = null;

function onLeadsFilter() {{
  const y = document.getElementById('lv-year')?.value;
  const m = document.getElementById('lv-month')?.value;
  if (!y || !m) return;
  const key = y + '-' + m;
  renderLeadsChart();
  renderUnitsChart(key);
  renderLeadsTable(key);
}}

function _metaDia(unidade, key) {{
  const metas = leadsAllMetas?.unidades?.[unidade];
  if (!metas) return null;
  const idx = leadsAllLabels.findIndex(l => {{
    const p = l.split('/'); return (p[1]+'-'+p[0]) === key;
  }});
  if (idx < 0) return null;
  const m = metas[idx];
  if (m == null) return null;
  const ratio = (key === leadsCurrentKey) ? leadsMetaRatio : 1.0;
  return Math.round(m * ratio);
}}

function _filtrarUnidades() {{
  if (!leadsOrdem) return [];
  if (leadsView === 'regiao') {{
    // leadsReg é 'Regiao 1' etc.; leadsRegiaoMap retorna 'Região 1' etc.
    const regNorm = leadsReg.replace('Regiao','Região').replace('Sao Paulo','São Paulo');
    return leadsOrdem.filter(u => leadsRegiaoMap[u] === regNorm);
  }}
  if (leadsView === 'unidade') {{
    return leadsOrdem.filter(u => u === leadsUnit);
  }}
  return leadsOrdem;
}}

function renderUnitsChart(key) {{
  if (typeof leadsAllData === 'undefined') return;

  const labelsU = [], dataLeads = [], dataMeta = [], colorsBar = [], colorsBorder = [];

  if (typeof leadsAllData !== 'undefined') {{
    const idx = leadsAllLabels.findIndex(l => {{ const p=l.split('/'); return (p[1]+'-'+p[0])===key; }});
    const unidadesFiltradas = _filtrarUnidades();
    unidadesFiltradas.forEach(u => {{
      const leads = idx >= 0 ? (leadsAllData.unidades?.[u]?.[idx] ?? 0) : 0;
      const meta  = _metaDia(u, key);
      labelsU.push(u);
      dataLeads.push(leads);
      dataMeta.push(meta);
      const ok = meta != null ? leads >= meta : null;
      colorsBar.push(ok === true ? 'rgba(46,125,50,0.70)' : ok === false ? 'rgba(198,40,40,0.70)' : 'rgba(158,158,158,0.45)');
      colorsBorder.push(ok === true ? '#2E7D32' : ok === false ? '#C62828' : '#9E9E9E');
    }});
  }}

  const C_ABOVE  = 'rgba(16,185,129,0.80)';  // emerald
  const C_BELOW  = 'rgba(239,68,68,0.78)';   // rose
  const C_NONE   = 'rgba(148,163,184,0.40)';
  const CB_ABOVE = '#059669';
  const CB_BELOW = '#DC2626';
  const CB_NONE  = '#94A3B8';

  const barsColor   = dataMeta.map((m,i) => m==null ? C_NONE  : dataLeads[i]>=m ? C_ABOVE  : C_BELOW);
  const barsBorder  = dataMeta.map((m,i) => m==null ? CB_NONE : dataLeads[i]>=m ? CB_ABOVE : CB_BELOW);

  const datasets = [
    {{ label: 'Leads', data: dataLeads,
       backgroundColor: barsColor, borderColor: barsBorder,
       borderWidth: 0, borderRadius: 5, borderSkipped: false,
       categoryPercentage: 0.65, barPercentage: 0.85, order: 1 }},
    {{ label: 'Meta/dia', data: dataMeta, type: 'line',
       borderColor: '#94A3B8', borderDash: [5,4], backgroundColor: 'transparent',
       borderWidth: 1.5, pointRadius: 0, tension: 0,
       fill: false, spanGaps: true, order: 0, datalabels: {{display:false}} }},
  ];

  if (unitsChart) {{
    unitsChart.data.labels   = labelsU;
    unitsChart.data.datasets = datasets;
    unitsChart.update('none');
  }} else {{
    const ctx = document.getElementById('chart-leads-units');
    if (!ctx) return;
    unitsChart = new Chart(ctx, {{
      type: 'bar',
      data: {{ labels: labelsU, datasets }},
      options: {{
        responsive: true,
        animation: {{ duration: 400, easing: 'easeOutQuart' }},
        plugins: {{
          legend: {{ display: false }},
          datalabels: {{
            display: ctx => ctx.datasetIndex === 0,
            anchor: 'end', align: 'top', offset: 2,
            font: {{ size: 10, weight: '600', family: 'inherit' }},
            color: ctx => barsBorder[ctx.dataIndex],
            formatter: v => v != null ? v.toLocaleString('pt-BR') : '',
            clip: false,
          }},
          tooltip: {{
            backgroundColor: 'rgba(15,23,42,0.88)',
            titleColor: '#F1F5F9',
            bodyColor: '#CBD5E1',
            padding: 10,
            cornerRadius: 8,
            boxPadding: 4,
            callbacks: {{
              title: ctx => ctx[0].label,
              label: ctx => {{
                const i = ctx.dataIndex;
                if (ctx.datasetIndex === 1) {{
                  const m = dataMeta[i];
                  return ' Meta/dia: ' + (m != null ? m.toLocaleString('pt-BR') : '—');
                }}
                const diff = dataMeta[i] != null ? dataLeads[i] - dataMeta[i] : null;
                const diffStr = diff != null
                  ? (diff >= 0 ? '  +' + diff.toLocaleString('pt-BR') : '  ' + diff.toLocaleString('pt-BR'))
                  : '';
                return [
                  ' Leads: ' + dataLeads[i].toLocaleString('pt-BR') + diffStr,
                  dataMeta[i] != null ? ' Meta/dia: ' + dataMeta[i].toLocaleString('pt-BR') : '',
                ].filter(Boolean);
              }},
            }}
          }},
        }},
        scales: {{
          x: {{
            offset: true,
            grid: {{ display: false }},
            border: {{ display: false }},
            ticks: {{ font: {{ size: 10, family: 'inherit' }}, maxRotation: 40, minRotation: 40,
                      color: '#94A3B8', padding: 4 }},
          }},
          y: {{
            beginAtZero: false,
            grid: {{ color: 'rgba(241,245,249,0.9)', drawBorder: false }},
            border: {{ display: false, dash: [4,4] }},
            ticks: {{
              font: {{ size: 10 }}, color: '#94A3B8', maxTicksLimit: 4, padding: 8,
              callback: v => v >= 1000 ? (v/1000).toLocaleString('pt-BR',{{maximumFractionDigits:1}})+'k' : v.toLocaleString('pt-BR'),
            }},
            suggestedMin: (() => {{
              const all = [...dataLeads, ...dataMeta].filter(v => v != null && v > 0);
              if (!all.length) return 0;
              const mn = Math.min(...all), mx = Math.max(...all);
              return Math.max(0, Math.floor(mn - (mx - mn) * 0.5));
            }})(),
          }},
        }},
        layout: {{ padding: {{ top: 22, right: 8, bottom: 0, left: 4 }} }},
      }},
    }});
  }}

  // Atualiza label do período
  const lbl = document.getElementById('lv-period-label');
  if (lbl) lbl.textContent = 'Mês: '+key.slice(5)+'/'+key.slice(0,4);
}}

function renderLeadsTable(key) {{
  if (typeof leadsAllData === 'undefined') return;
  const idx = leadsAllLabels.findIndex(l => {{ const p=l.split('/'); return (p[1]+'-'+p[0])===key; }});
  const RNAMES = {{'Regiao 1':'Região 1','Regiao 2':'Região 2','Regiao 3':'Região 3','Sao Paulo':'São Paulo'}};

  const rows = _filtrarUnidades().map(u => {{
    const leads = idx >= 0 ? (leadsAllData.unidades?.[u]?.[idx] ?? 0) : 0;
    const meta  = _metaDia(u, key);
    const falta = meta != null ? leads - meta : null;
    const conv  = (typeof leadsConversao !== 'undefined') ? (leadsConversao[u]?.[key] ?? null) : null;
    const icl   = (conv != null && leads > 0) ? (conv / leads * 100) : null;
    return {{ u, leads, meta, falta, conv, icl }};
  }}).sort((a,b) => b.leads - a.leads);

  // Obtém regiões do JSON de unidades
  const getRegiao = u => {{
    for (const r of ['Regiao 1','Regiao 2','Regiao 3','Sao Paulo']) {{
      const d = leadsAllData.regioes?.[r];
      // não temos mapeamento direto — vamos usar leadsOrdem (injected)
    }}
    return '';
  }};

  let html = '';
  let totLeads = 0, totMeta = 0, totFalta = 0, totConv = 0;
  let hasMeta = false, hasConv = false;

  rows.forEach((r, i) => {{
    totLeads += r.leads;
    if (r.meta != null) {{ totMeta += r.meta; hasMeta = true; }}
    if (r.falta != null) totFalta += r.falta;
    if (r.conv  != null) {{ totConv  += r.conv;  hasConv  = true; }}

    const faltaStr  = r.falta  != null ? (r.falta >= 0 ? '<span style="color:#2E7D32">+'+r.falta.toLocaleString('pt-BR')+'</span>' : '<span style="color:#C62828">'+r.falta.toLocaleString('pt-BR')+'</span>') : '—';
    const metaStr   = r.meta   != null ? r.meta.toLocaleString('pt-BR') : '—';
    const iclStr    = r.icl    != null ? r.icl.toFixed(1)+'%' : '—';
    const iclColor  = r.icl    != null ? (r.icl >= 30 ? '#2E7D32' : r.icl >= 15 ? '#F57C00' : '#C62828') : 'var(--subtexto)';

    html += `<tr>
      <td class="rank">#${{i+1}}</td>
      <td class="td-nome">${{r.u}}</td>
      <td style="color:var(--subtexto);font-size:12px">${{leadsRegiaoMap[r.u]||''}}</td>
      <td class="num blue-txt">${{r.leads.toLocaleString('pt-BR')}}</td>
      <td class="num">${{metaStr}}</td>
      <td class="num">${{faltaStr}}</td>
      <td class="num" style="color:var(--subtexto)">${{r.conv != null ? r.conv.toLocaleString('pt-BR') : '—'}}</td>
      <td class="num" style="color:${{iclColor}};font-weight:700">${{iclStr}}</td>
    </tr>`;
  }});

  const totFaltaStr = hasMeta ? (totFalta >= 0 ? '<span style="color:#2E7D32">+'+totFalta.toLocaleString('pt-BR')+'</span>' : '<span style="color:#C62828">'+totFalta.toLocaleString('pt-BR')+'</span>') : '—';
  const totIcl = (hasConv && totLeads > 0) ? (totConv/totLeads*100).toFixed(1)+'%' : '—';
  const footHtml = `<tr style="background:var(--cinza2);font-weight:700">
    <td colspan="3" class="td-nome">TOTAL REDE</td>
    <td class="num blue-txt">${{totLeads.toLocaleString('pt-BR')}}</td>
    <td class="num">${{hasMeta ? totMeta.toLocaleString('pt-BR') : '—'}}</td>
    <td class="num">${{totFaltaStr}}</td>
    <td class="num" style="color:var(--subtexto)">${{hasConv ? totConv.toLocaleString('pt-BR') : '—'}}</td>
    <td class="num">${{totIcl}}</td>
  </tr>`;

  const tbody = document.getElementById('tbl-leads-body');
  const tfoot = document.getElementById('tbl-leads-foot');
  if (tbody) tbody.innerHTML = html;
  if (tfoot) tfoot.innerHTML = footHtml;

  const sub = document.getElementById('tbl-leads-subtitle');
  if (sub) sub.textContent = key.slice(5)+'/'+key.slice(0,4)+' · ordenado por volume';
}}

// Mapa unidade → região (injetado do Python)
const leadsRegiaoMap = {_json.dumps({nome: info.get('regiao','').replace('Regiao','Região') for nome, info in unidades.items()})};

function initLeadsFilter() {{
  const y = document.getElementById('lv-year');
  const m = document.getElementById('lv-month');
  if (!y || !m) return;
  // Seta mês/ano atual como padrão
  const [curY, curM] = leadsCurrentKey.split('-');
  if (curY) y.value = curY;
  if (curM) m.value = curM;
  onLeadsFilter();
}}

// ── Navigation ─────────────────────────────────────────────────
function showTab(id, tabId) {{
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  const tab = typeof tabId === 'string' ? document.getElementById(tabId) : tabId;
  if (tab) tab.classList.add('active');
  window.scrollTo({{ top: 0, behavior: 'smooth' }});
  if (id === 'instagram' && !igChart) updateIgChart();
  if (id === 'leads') {{ if (!leadsChart) initLeadsChart(); initLeadsFilter(); }}
}}

document.addEventListener('DOMContentLoaded', () => {{
  initMiniChart();
  updateIgChart();
}});
</script>"""

    # ── HTML COMPLETO ──────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Greenlife — Dashboard</title>
<style>
  :root {{
    --verde:      #2E7D32;
    --verde-pale: #E8F5E9;
    --cinza:      #F0F2F5;
    --cinza2:     #F8F9FA;
    --texto:      #1A1A1A;
    --subtexto:   #6B7280;
    --borda:      #E5E7EB;
    --branco:     #FFFFFF;
    --laranja:    #F57C00;
    --azul:       #1565C0;
    --azul-pale:  #E3F2FD;
    --roxo:       #6A1B9A;
    --roxo-pale:  #F3E5F5;
  }}
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
          background:var(--cinza); color:var(--texto); min-height:100vh; }}

  /* HEADER */
  .topbar {{ background:#1B3A2F; padding:0 32px; display:flex; align-items:center;
             gap:24px; height:60px; position:sticky; top:0; z-index:100;
             box-shadow:0 2px 8px rgba(0,0,0,.18); }}
  .logo-mark {{ width:34px; height:34px; background:#4CAF50; border-radius:8px;
                display:flex; align-items:center; justify-content:center;
                color:white; font-weight:800; font-size:16px; flex-shrink:0; }}
  .topbar-title {{ font-size:16px; font-weight:700; color:#fff; }}
  .update-stamp {{ display:flex; align-items:center; gap:14px; margin-left:16px; }}
  .update-item {{ display:flex; align-items:center; gap:5px; font-size:11px; color:rgba(255,255,255,.5); white-space:nowrap; }}
  .update-dot {{ width:6px; height:6px; border-radius:50%; background:rgba(255,255,255,.3); flex-shrink:0; }}
  .update-dot.ok {{ background:#4CAF50; }}
  .tabs {{ margin-left:auto; display:flex; gap:4px; }}
  .tab {{ padding:6px 18px; border-radius:8px; font-size:14px; font-weight:500;
          cursor:pointer; color:rgba(255,255,255,.65); background:none; border:none;
          transition:background .15s,color .15s; }}
  .tab:hover {{ background:rgba(255,255,255,.12); color:#fff; }}
  .tab.active {{ background:#4CAF50; color:white; }}

  /* BACK BAR */
  .back-bar {{ display:flex; align-items:center; gap:8px; padding:10px 0 4px;
               margin-bottom:8px; }}
  .back-btn {{ background:var(--cinza2); border:1px solid var(--borda); border-radius:8px;
               padding:5px 12px; font-size:13px; font-weight:600; color:var(--texto);
               cursor:pointer; transition:background .15s; }}
  .back-btn:hover {{ background:var(--borda); }}
  .back-sep {{ color:var(--subtexto); }}
  .back-label {{ font-size:13px; font-weight:600; color:var(--subtexto); }}

  /* CONTENT */
  .page {{ display:none; padding:20px 28px; }}
  .page.active {{ display:block; }}

  /* SECTION LABEL */
  .section-label {{ font-size:13px; font-weight:700; text-transform:uppercase;
                    letter-spacing:.8px; color:var(--texto); margin-bottom:10px;
                    padding:7px 14px; background:var(--branco); border:1px solid var(--borda);
                    border-left:4px solid var(--verde); border-radius:0 8px 8px 0;
                    box-shadow:0 1px 3px rgba(0,0,0,.04); }}

  /* VER MAIS */
  .ver-mais-btn {{ border:none; border-radius:20px; padding:5px 14px;
                   font-size:12px; font-weight:600; cursor:pointer; transition:opacity .15s; }}
  .ver-mais-btn:hover {{ opacity:.75; }}
  .ver-mais-blue   {{ background:var(--azul-pale);  color:var(--azul); }}
  .ver-mais-purple {{ background:var(--roxo-pale);  color:var(--roxo); }}
  .ver-mais-green  {{ background:var(--verde-pale); color:var(--verde); }}

  /* BIG CARDS */
  .big-cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr));
                gap:10px; margin-bottom:8px; }}
  .big-card {{ background:var(--branco); border-radius:12px; padding:14px 18px;
               border:1px solid var(--borda); box-shadow:0 1px 4px rgba(0,0,0,.05); }}
  .big-card.green  {{ border-top:3px solid var(--verde); }}
  .big-card.orange {{ border-top:3px solid var(--laranja); }}
  .big-card.blue   {{ border-top:3px solid var(--azul); }}
  .big-card.purple {{ border-top:3px solid var(--roxo); }}
  .bc-label {{ font-size:11px; color:var(--subtexto); text-transform:uppercase;
               letter-spacing:.5px; margin-bottom:6px; }}
  .bc-valor {{ font-size:28px; font-weight:800; line-height:1; }}
  .big-card.green  .bc-valor {{ color:var(--verde); }}
  .big-card.orange .bc-valor {{ color:var(--laranja); }}
  .big-card.blue   .bc-valor {{ color:var(--azul); }}
  .big-card.purple .bc-valor {{ color:var(--roxo); }}
  .bc-sub {{ font-size:12px; color:var(--subtexto); margin-top:8px; }}
  .bc-detail {{ margin-top:14px; padding-top:12px; border-top:1px solid var(--borda); }}
  .bc-detail-row {{ display:flex; justify-content:space-between; font-size:12px;
                    color:var(--subtexto); padding:3px 0; }}
  .bc-detail-row span:last-child {{ font-weight:600; color:var(--texto); }}

  /* PANEL */
  .panel {{ background:var(--branco); border-radius:14px; border:1px solid var(--borda);
            box-shadow:0 1px 4px rgba(0,0,0,.05); overflow:hidden; margin-bottom:20px; }}
  .panel-header {{ padding:16px 24px; border-bottom:1px solid var(--borda);
                   display:flex; align-items:center; gap:10px; }}
  .panel-header h2 {{ font-size:15px; font-weight:700; }}
  .panel-header p  {{ font-size:12px; color:var(--subtexto); margin-top:2px; }}
  .dot {{ width:8px; height:8px; border-radius:50%; background:var(--verde); flex-shrink:0; }}
  .blue-dot {{ background:var(--azul); }}

  /* TABLE */
  .table-wrap {{ overflow-x:auto; }}
  table {{ width:100%; border-collapse:collapse; }}
  th {{ font-size:11px; text-transform:uppercase; letter-spacing:.5px; color:var(--subtexto);
        padding:10px 16px; text-align:left; background:var(--cinza2); white-space:nowrap; }}
  th.num, td.num {{ text-align:right; }}
  td {{ padding:12px 16px; font-size:14px; border-top:1px solid var(--borda); white-space:nowrap; }}
  tr:hover td {{ background:var(--cinza2); }}
  td.rank    {{ color:var(--subtexto); font-size:12px; width:36px; }}
  td.td-nome {{ font-weight:600; }}
  .green-txt {{ color:var(--verde); font-weight:700; }}
  .blue-txt  {{ color:var(--azul);  font-weight:700; }}
  .badge {{ background:var(--verde-pale); color:var(--verde); padding:2px 10px;
             border-radius:20px; font-size:12px; font-weight:600; }}
  .empty {{ padding:40px; text-align:center; color:var(--subtexto); font-size:14px; }}

  /* INFO BADGE */
  .info-badge {{
    display:inline-flex; align-items:center; justify-content:center;
    width:13px; height:13px; border-radius:50%;
    background:var(--subtexto); color:#fff;
    font-size:8px; font-weight:700; font-style:italic;
    vertical-align:middle; opacity:.55; cursor:default;
    margin-left:3px; flex-shrink:0;
    transition:opacity .15s;
  }}
  .info-badge:hover {{ opacity:1; }}

  /* SORTABLE TH */
  .th-sort {{ cursor:pointer; user-select:none; transition:background .1s; position:relative; padding-right:22px; }}
  .th-sort:hover {{ background:var(--cinza); color:var(--texto); }}
  .th-sort::after {{ content:'⇅'; position:absolute; right:6px; top:50%; transform:translateY(-50%);
                     font-size:9px; opacity:.35; font-weight:400; }}
  .th-sort.s-asc::after  {{ content:'▲'; opacity:.8; color:var(--azul); }}
  .th-sort.s-desc::after {{ content:'▼'; opacity:.8; color:var(--azul); }}

  /* VIEW BUTTONS */
  .view-btns {{ display:flex; gap:4px; }}
  .pill {{ font-size:12px; font-weight:600; padding:4px 12px; border-radius:20px;
           border:1px solid var(--borda); background:var(--cinza2); color:var(--subtexto);
           cursor:pointer; transition:all .15s; white-space:nowrap; }}
  .pill:hover {{ background:var(--borda); color:var(--texto); }}
  .pill.active {{ background:var(--verde); color:white; border-color:var(--verde); }}

  /* LAYOUT */
  .two-col {{ display:grid; grid-template-columns:1fr 1fr; gap:20px; }}
  .period-bar {{ font-size:13px; color:var(--subtexto); margin-bottom:16px;
                 display:flex; align-items:center; gap:8px; }}
  .period-badge {{ background:var(--azul-pale); color:var(--azul); padding:2px 10px;
                   border-radius:20px; font-size:12px; font-weight:600; }}

  footer {{ text-align:center; padding:24px; font-size:12px; color:var(--subtexto); }}

  @media (max-width:900px) {{
    .two-col {{ grid-template-columns:1fr; }}
    .tabs {{ margin-left:0; }}
    .topbar {{ flex-wrap:wrap; height:auto; padding:12px 16px; gap:12px; }}
    .page {{ padding:16px; }}
  }}
</style>
</head>
<body>

<div class="topbar">
  <div class="logo-mark" onclick="showTab('geral',document.getElementById('tab-geral'))" style="cursor:pointer">GL</div>
  <div class="topbar-title" onclick="showTab('geral',document.getElementById('tab-geral'))" style="cursor:pointer">Greenlife Dashboard</div>
  <div class="update-stamp">
    <div class="update-item"><div class="update-dot ok"></div>Leads {gerado_em}</div>
    <div class="update-item"><div class="update-dot ok"></div>Instagram {_data_fim.strftime('%d/%m/%Y')}</div>
    <div class="update-item"><div class="update-dot ok"></div>Clube {resumo['data']}</div>
  </div>
  <div class="tabs">
    <button class="tab active" id="tab-geral"      onclick="showTab('geral',this)">Visão Geral</button>
    <button class="tab"        id="tab-instagram"  onclick="showTab('instagram',this)">Instagram</button>
    <button class="tab"        id="tab-parceiros"  onclick="showTab('parceiros',this)">Parceiros</button>
    <button class="tab"        id="tab-leads"      onclick="showTab('leads',this)">Leads Pacto</button>
  </div>
</div>

<div id="geral"      class="page active">{aba_geral}</div>
<div id="instagram"  class="page">{aba_instagram}</div>
<div id="parceiros"  class="page">{aba_parceiros}</div>
<div id="leads"      class="page">{aba_leads}</div>

<footer>Greenlife Academias · {datetime.now().strftime('%d/%m/%Y %H:%M')}</footer>

{chart_script}
</body>
</html>"""

    Path(saida).write_text(html, encoding="utf-8")
    print(f"[OK] Dashboard gerado: {Path(saida).resolve()}")


def salvar_resumo_macro(arquivo_excel="dados_greenlife.xlsx", saida="resumo_macro.xlsx"):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(arquivo_excel)

    ws_r = wb["Resumo Diario"]
    snap_mensal = {}
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            partes = str(row[0])[:10].split("/")
            if len(partes) != 3:
                continue
            mk = f"{partes[1]}/{partes[2]}"
            snap_mensal[mk] = {
                "parceiros": row[4] if row[4] not in (None,"-","") else None,
                "promocoes": row[1] if row[1] not in (None,"-","") else None,
            }
        except Exception:
            pass

    ws_st = wb["Serie Temporal"]
    cupons_mensal = {}
    for row in ws_st.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        try:
            partes = str(row[0]).split("/")
            if len(partes) != 3 or len(partes[2]) != 4:
                continue
            mk = f"{partes[1]}/{partes[2]}"
            g = int(row[1]) if row[1] else 0
            u = int(row[2]) if row[2] else 0
            if mk not in cupons_mensal:
                cupons_mensal[mk] = {"gerados": 0, "utilizados": 0}
            cupons_mensal[mk]["gerados"]    += g
            cupons_mensal[mk]["utilizados"] += u
        except Exception:
            pass

    ig_seg_mensal, ig_int_mensal = {}, {}
    arq = INSTAGRAM_DIR / "base_seguidores.xlsx"
    if arq.exists():
        ws = openpyxl.load_workbook(arq).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1] or not row[3]:
                continue
            try:
                partes = str(row[3]).split("/")
                mk = f"{partes[1]}/{partes[2]}"
                ig_seg_mensal.setdefault(mk, {})[row[1]] = row[4]
            except Exception:
                pass

    arq = INSTAGRAM_DIR / "base_interacoes_total.xlsx"
    if arq.exists():
        ws = openpyxl.load_workbook(arq).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1] or not row[3]:
                continue
            try:
                partes = str(row[3]).split("/")
                mk = f"{partes[1]}/{partes[2]}"
                ig_int_mensal.setdefault(mk, {})[row[1]] = row[4]
            except Exception:
                pass

    headers = ["Mês","Parceiros","Promoções","Cupons Gerados","Cupons Utilizados",
               "Taxa Utilização","Seguidores Total","Interações Total"]
    COR_H = "1A3A5C"
    thin  = Side(style="thin", color="CCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    saida_path = Path(saida)
    if saida_path.exists():
        wbo = openpyxl.load_workbook(saida)
        ws  = wbo.active
        meses_existentes = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    else:
        wbo = openpyxl.Workbook()
        ws  = wbo.active
        ws.title = "Resumo Macro"
        ws.freeze_panes = "A2"
        meses_existentes = set()
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=COR_H)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
        ws.row_dimensions[1].height = 32

    todos_meses = sorted(
        set(snap_mensal) | set(cupons_mensal) | set(ig_seg_mensal),
        key=_mes_sort
    )
    meses_novos = [mk for mk in todos_meses if mk not in meses_existentes]

    if not meses_novos:
        print("[OK] Resumo macro já está atualizado.")
        return

    for r, mk in enumerate(meses_novos, ws.max_row + 1):
        zebra = r % 2 == 0
        fill  = PatternFill("solid", fgColor="EFF4FB") if zebra else None
        snap   = snap_mensal.get(mk, {})
        cupons = cupons_mensal.get(mk, {})
        segs   = ig_seg_mensal.get(mk, {})
        ints   = ig_int_mensal.get(mk, {})
        g = cupons.get("gerados", 0)
        u = cupons.get("utilizados", 0)
        taxa = round(u / g * 100, 1) if g else None
        seg_total = sum(v for v in segs.values() if v) or None
        int_total = sum(v for v in ints.values() if v) or None
        vals = [mk, snap.get("parceiros"), snap.get("promocoes"),
                g or None, u or None, taxa, seg_total, int_total]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
            if fill:
                c.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)

    wbo.save(saida)
    print(f"[OK] Resumo macro: {len(meses_novos)} mes(es) novo(s) em {Path(saida).resolve()}")


if __name__ == "__main__":
    _base = Path(__file__).parent
    gerar_dashboard(
        arquivo_excel=str(_base / "dados_greenlife.xlsx"),
        saida=str(_base / "dashboard.html"),
    )
    salvar_resumo_macro(
        arquivo_excel=str(_base / "dados_greenlife.xlsx"),
        saida=str(_base / "resumo_macro.xlsx"),
    )
