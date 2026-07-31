"""
Salva snapshot diário de leads por unidade no Excel.
Colunas: Unidade | Leads no mês | Data de referência
Executa após coletar_leads.py — lê leads_dados.json.
"""

import json
import os
from datetime import datetime, timedelta, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(__file__)
LEADS_JSON = os.path.join(DIR, "leads_dados.json")
OUTPUT = os.path.join(DIR, "leads_diario.xlsx")

BRT = timezone(timedelta(hours=-3))
HEADERS = ["Unidade", "Leads no mês", "Data de referência"]

COR_HEADER = "1A3A5C"
COR_ZEBRA = "EFF4FB"
COR_BORDA = "CCCCCC"


def _thin():
    return Side(style="thin", color=COR_BORDA)


def _header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    b = _thin()
    cell.border = Border(left=b, right=b, top=b, bottom=b)


def _dado(cell, zebra=False):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if zebra:
        cell.fill = PatternFill("solid", fgColor=COR_ZEBRA)
    b = _thin()
    cell.border = Border(left=b, right=b, top=b, bottom=b)


def _ajusta(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 35)


def carregar_ou_criar():
    if os.path.exists(OUTPUT):
        wb = openpyxl.load_workbook(OUTPUT)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads Diário"
        ws.freeze_panes = "A2"
        for i, h in enumerate(HEADERS, 1):
            _header(ws.cell(row=1, column=i, value=h))
        ws.row_dimensions[1].height = 30
    return wb, ws


def ja_existe(ws, unidade, data_str):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == unidade and str(row[2]) == data_str:
            return True
    return False


def main():
    if not os.path.exists(LEADS_JSON):
        print("leads_dados.json não encontrado. Rode coletar_leads.py primeiro.")
        return

    with open(LEADS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    hoje = datetime.now(BRT)
    data_ref = hoje.strftime("%d/%m/%Y")
    mes_key = hoje.strftime("%Y-%m")

    unidades = data.get("unidades", {})
    wb, ws = carregar_ou_criar()

    novos = 0
    for unidade, info in sorted(unidades.items()):
        leads = info.get("meses", {}).get(mes_key, 0) or 0
        if ja_existe(ws, unidade, data_ref):
            print(f"  [{unidade}] já registrado em {data_ref}, pulando.")
            continue
        n = ws.max_row + 1
        zebra = n % 2 == 0
        for i, val in enumerate([unidade, leads, data_ref], 1):
            _dado(ws.cell(row=n, column=i, value=val), zebra)
        novos += 1

    if novos == 0:
        print(f"Nenhum dado novo para {data_ref}.")
        return

    _ajusta(ws)
    wb.save(OUTPUT)
    print(f"Salvo: {OUTPUT}  ({novos} unidades em {data_ref})")


if __name__ == "__main__":
    main()
