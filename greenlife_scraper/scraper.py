# -*- coding: utf-8 -*-
"""
Greenlife Scraper — acessa a API REST do backend diretamente.
Fluxo: login via Playwright -> captura JWT -> chama API com requests.
"""
import asyncio
import os
import sys
import requests
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EMAIL    = os.environ.get("PACTO_EMAIL") or ""
SENHA    = os.environ.get("PACTO_SENHA") or ""
URL_SITE = "https://parceirogreenlife.com.br"
API_BASE = "http://greenlife-backend-v2.deway.com.br/api"
ARQUIVO  = "dados_greenlife.xlsx"
VERDE    = "1B5E20"


def log(msg):
    print(msg.encode("ascii", "replace").decode("ascii"))
    sys.stdout.flush()


# ─── excel ───────────────────────────────────────────────────────────────────

def cabecalho_sheet(ws, colunas):
    ws.append([c[0] for c in colunas])
    for i, cell in enumerate(ws[1]):
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=VERDE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, (_, w) in enumerate(colunas, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 20


# ─── extração de cards via JS ────────────────────────────────────────────────

async def extrair_cards_js(page) -> dict:
    return await page.evaluate("""
        () => {
            const pares = {};
            Array.from(document.querySelectorAll('*')).forEach(el => {
                const txt = (el.innerText || '').trim();
                const linhas = txt.split('\\n').map(l => l.trim()).filter(Boolean);
                if (linhas.length >= 2 && linhas.length <= 6) {
                    linhas.forEach((linha, i) => {
                        if (/^\\d+$/.test(linha)) {
                            const label = linhas[i - 1] || linhas[i + 1] || '';
                            if (label.length > 3) pares[label.toLowerCase()] = linha;
                        }
                    });
                }
            });
            return pares;
        }
    """)


# ─── step 1: login via playwright e captura o JWT ────────────────────────────

async def obter_token() -> tuple:
    log("[1/4] Fazendo login e capturando token JWT...")
    token = None

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()

        async def on_response(response):
            nonlocal token
            if "doapicallfromserver" in response.url and response.status == 200:
                try:
                    body = await response.json()
                    props = body.get("properties", {})
                    t = props.get("headers_Authorization", "")
                    if t.startswith("Bearer ") and not token:
                        token = t.replace("Bearer ", "").strip()
                except Exception:
                    pass

        page.on("response", on_response)

        await page.goto(f"{URL_SITE}/auth", wait_until="networkidle")
        await page.wait_for_timeout(2000)
        await page.locator('input[type="email"]').fill(EMAIL)
        await page.locator('input[type="password"]').fill(SENHA)
        await page.locator('button:has-text("Entrar")').click()

        for _ in range(30):
            await page.wait_for_timeout(500)
            if "/auth" not in page.url:
                break
        else:
            raise RuntimeError("Login falhou.")

        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)

        # navega para Promocoes para disparar a chamada de API com token
        if not token:
            await page.evaluate("""
                () => {
                    const els = Array.from(document.querySelectorAll('a, button, span'));
                    const el = els.find(e => e.innerText.trim().toLowerCase().includes('promo'));
                    if (el) el.click();
                }
            """)
            await page.wait_for_timeout(4000)

        # captura cupons e totais de parceiros dos cards da pagina
        cards = await extrair_cards_js(page)
        log(f"  -> Cards dashboard: {cards}")

        await page.evaluate("""
            () => {
                const els = Array.from(document.querySelectorAll('a, button, span'));
                const el = els.find(e => e.innerText.trim().toLowerCase().includes('parceiros'));
                if (el) el.click();
            }
        """)
        await page.wait_for_timeout(3000)
        cards_parc = await extrair_cards_js(page)
        log(f"  -> Cards parceiros: {cards_parc}")

        await browser.close()

    if not token:
        raise RuntimeError("Token JWT nao capturado.")

    log(f"[OK] Token capturado: {token[:30]}...")

    cupons = {
        "cupons_gerados":    cards.get("cupons gerados", "-"),
        "cupons_utilizados": cards.get("cupons utilizados", "-"),
    }
    def buscar_card(cards, *partes):
        for k, v in cards.items():
            if all(p in k for p in partes):
                return v
        return "-"

    parc_resumo = {
        "com_usuario": buscar_card(cards_parc, "parceiros", "com"),
        "sem_usuario": buscar_card(cards_parc, "parceiros", "sem"),
    }
    return token, cupons, parc_resumo


# ─── step 2: chama a API diretamente ─────────────────────────────────────────

def api_get(endpoint: str, token: str, params: dict = None) -> dict:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
    }
    url = f"{API_BASE}/{endpoint}"
    r = requests.get(url, headers=headers, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def buscar_todas_paginas(endpoint: str, token: str, max_retries: int = 3) -> list:
    """Busca todas as páginas de um endpoint paginado, com retry por página."""
    import time
    resultados = []
    pagina = 1

    while True:
        log(f"  -> Pagina {pagina}...")
        data = None
        for tentativa in range(1, max_retries + 1):
            try:
                data = api_get(endpoint, token, params={"page": pagina})
                break
            except Exception as e:
                log(f"     [tentativa {tentativa}/{max_retries}] erro: {e}")
                if tentativa < max_retries:
                    time.sleep(5 * tentativa)
        if data is None:
            log(f"     [AVISO] pagina {pagina} falhou apos {max_retries} tentativas — interrompendo paginacao.")
            break

        if "pagination" in data:
            items = data["pagination"].get("results", [])
            proximo = data["pagination"].get("next")
        elif "results" in data:
            items = data["results"]
            proximo = data.get("next")
        else:
            items = data if isinstance(data, list) else []
            proximo = None

        resultados.extend(items)
        log(f"     {len(items)} items (total ate agora: {len(resultados)})")

        if not proximo or not items:
            break
        pagina += 1

    return resultados


# ─── step 3: busca dados do dashboard ────────────────────────────────────────

def capturar_dashboard_api(token: str) -> dict:
    log("\n[2/5] Dashboard — get_card_data (API)...")
    card = api_get("dashboard/get_card_data", token)
    log(f"  -> {card}")
    return {
        "total_promocoes":   str(card.get("count_promotions", "-")),
        "cupons_gerados":    str(card.get("count_coupons", "-")),
        "cupons_utilizados": str(card.get("count_coupons_used", "-")),
    }


def capturar_serie_temporal(token: str) -> list:
    """Retorna lista de dicts {date, coupons_genereted, coupons_used} do historico completo."""
    log("\n[3/5] Serie temporal diaria (API)...")
    dados = api_get("dashboard/get_coupons_used", token)
    if isinstance(dados, list):
        log(f"  -> {len(dados)} dias de historico (de {dados[0]['date']} a {dados[-1]['date']})")
        return dados
    return []


# ─── step 5: promoções ───────────────────────────────────────────────────────

def capturar_promocoes_api(token: str) -> tuple:
    log("\n[4/5] Promocoes (API)...")
    items = buscar_todas_paginas("dashboard/", token)
    log(f"  -> Total: {len(items)} promocoes")

    vistos = set()
    linhas = []
    for p in items:
        pid = p.get("id")
        if pid in vistos:
            continue
        vistos.add(pid)
        situacao = "Ativo" if p.get("is_active") else "Inativo"
        tipo     = p.get("discount_type", "")
        valor    = p.get("discount_value", "")
        desconto = f"{valor}%" if tipo == "PERCENTAGE" else f"R$ {valor}" if valor else "-"
        max_u    = p.get("max_users", -1)
        limite   = "-" if max_u == -1 else str(max_u)
        expira   = p.get("expirationDate") or "Sem expiracao"
        linhas.append([
            p.get("name", ""),
            str(pid),
            desconto,
            situacao,
            limite,
            str(p.get("num_coupons", 0)),
            str(p.get("num_coupons_used", 0)),
            "-",
            expira,
            str(p.get("partner", "")),
        ])

    ativas   = sum(1 for p in items if p.get("is_active"))
    inativas = len(items) - ativas
    resumo = {"total": str(len(items)), "ativas": str(ativas), "inativas": str(inativas)}
    log(f"  -> {resumo}")
    return resumo, linhas


# ─── step 6: parceiros ───────────────────────────────────────────────────────

def capturar_parceiros_api(token: str, parc_resumo_page: dict) -> tuple:
    log("\n[5/5] Parceiros (API)...")
    items = buscar_todas_paginas("dashboard/list_partners", token)
    log(f"  -> Total: {len(items)} parceiros")

    vistos = set()
    linhas = []
    com_usuario = 0
    for p in items:
        pid = p.get("id")
        if pid in vistos:
            continue
        vistos.add(pid)
        usuarios = p.get("partner_owned_user", [])
        if usuarios:
            u           = usuarios[0]
            responsavel = u.get("first_name", "")
            email       = u.get("email", "")
            com_usuario += 1
        else:
            responsavel = "Nao cadastrado"
            email       = ""
        cat    = p.get("category") or {}
        bairro = p.get("neighborhood") or p.get("neighrborhood") or "-"
        linhas.append([
            str(p.get("id", "")),
            p.get("name", ""),
            f"{responsavel} / {email}".strip(" /") if email else responsavel,
            cat.get("name", "-"),
            bairro,
        ])

    resumo = {
        "total":       str(len(items)),
        "com_usuario": parc_resumo_page.get("com_usuario", str(com_usuario)),
        "sem_usuario": parc_resumo_page.get("sem_usuario", str(len(items) - com_usuario)),
    }
    log(f"  -> {resumo}")
    return resumo, linhas


# ─── excel ───────────────────────────────────────────────────────────────────

def salvar_excel(dashboard, promo_resumo, promo_linhas, parc_resumo, parc_linhas, serie_temporal):
    arquivo = Path(ARQUIVO)
    agora   = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.load_workbook(arquivo) if arquivo.exists() else openpyxl.Workbook()

    # limpa abas antigas com encoding quebrado
    for nome in list(wb.sheetnames):
        if nome not in ["Resumo Diario", "Promocoes", "Parceiros"]:
            del wb[nome]
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Resumo Diario (acumula historico)
    if "Resumo Diario" not in wb.sheetnames:
        ws = wb.create_sheet("Resumo Diario", 0)
        cabecalho_sheet(ws, [
            ("Data", 20), ("Total Promocoes", 20), ("Cupons Gerados", 18),
            ("Cupons Utilizados", 20), ("Total Parceiros", 20),
            ("Parceiros c/ Usuario", 22), ("Parceiros s/ Usuario", 22),
        ])
    else:
        ws = wb["Resumo Diario"]

    ws.append([
        agora,
        dashboard.get("total_promocoes", "-"),
        dashboard.get("cupons_gerados", "-"),
        dashboard.get("cupons_utilizados", "-"),
        parc_resumo.get("total", "-"),
        parc_resumo.get("com_usuario", "-"),
        parc_resumo.get("sem_usuario", "-"),
    ])

    # Promocoes (substituida a cada execucao)
    if "Promocoes" in wb.sheetnames:
        del wb["Promocoes"]
    ws_p = wb.create_sheet("Promocoes")
    cabecalho_sheet(ws_p, [
        ("Promocao", 60), ("ID", 10), ("Desconto", 12), ("Situacao", 12),
        ("Limite Cupons", 16), ("Cupons Gerados", 16), ("Cupons Utilizados", 18),
        ("Cupons Restantes", 18), ("Valido ate", 18), ("ID Parceiro", 12),
    ])
    for l in promo_linhas:
        ws_p.append(l)
    ws_p.append([])
    ws_p.append([f"Atualizado: {agora} | Ativas: {promo_resumo.get('ativas')} | Inativas: {promo_resumo.get('inativas')}"])

    # Parceiros (substituida a cada execucao)
    if "Parceiros" in wb.sheetnames:
        del wb["Parceiros"]
    ws_parc = wb.create_sheet("Parceiros")
    cabecalho_sheet(ws_parc, [
        ("Codigo", 10), ("Parceiro", 35), ("Responsavel / Email", 45),
        ("Categoria", 20), ("Bairro", 20),
    ])
    for l in parc_linhas:
        ws_parc.append(l)
    ws_parc.append([])
    ws_parc.append([f"Atualizado: {agora} | Total: {parc_resumo.get('total')} | c/ usuario: {parc_resumo.get('com_usuario')} | s/ usuario: {parc_resumo.get('sem_usuario')}"])

    # Serie Temporal (substituida a cada execucao)
    if "Serie Temporal" in wb.sheetnames:
        del wb["Serie Temporal"]
    ws_st = wb.create_sheet("Serie Temporal")
    cabecalho_sheet(ws_st, [
        ("Data", 14), ("Cupons Gerados no Dia", 22), ("Cupons Utilizados no Dia", 24),
    ])
    for d in serie_temporal:
        ws_st.append([
            d.get("date", ""),
            d.get("coupons_genereted", 0),
            d.get("coupons_used", 0),
        ])
    ws_st.append([])
    ws_st.append([f"Atualizado: {agora} | Total dias: {len(serie_temporal)}"])

    wb.save(arquivo)
    log(f"\n[OK] Arquivo salvo: {arquivo.resolve()}")


# ─── main ────────────────────────────────────────────────────────────────────

async def main():
    import traceback
    try:
        token, _, parc_resumo_page = await obter_token()
    except Exception as e:
        log(f"\n[ERRO] Login/token falhou: {e}")
        log(traceback.format_exc())
        return

    # Coleta dashboard + série temporal (críticos — sem estes não salvamos nada)
    try:
        dashboard      = capturar_dashboard_api(token)
        serie_temporal = capturar_serie_temporal(token)
    except Exception as e:
        log(f"\n[ERRO] Falha nos dados principais: {e}")
        log(traceback.format_exc())
        return

    # Promoções — não crítico
    try:
        promo_resumo, promo_linhas = capturar_promocoes_api(token)
    except Exception as e:
        log(f"\n[AVISO] Promocoes falharam: {e} — continuando sem elas.")
        promo_resumo = {"total": "-", "ativas": "-", "inativas": "-"}
        promo_linhas = []

    # Parceiros — não crítico (paginação pode falhar)
    try:
        parc_resumo, parc_linhas = capturar_parceiros_api(token, parc_resumo_page)
    except Exception as e:
        log(f"\n[AVISO] Parceiros falharam: {e} — salvando o restante sem parceiros.")
        parc_resumo = {"total": "-", "com_usuario": "-", "sem_usuario": "-"}
        parc_linhas = []

    salvar_excel(dashboard, promo_resumo, promo_linhas, parc_resumo, parc_linhas, serie_temporal)

    try:
        from gerar_dashboard import gerar_dashboard
        gerar_dashboard()
    except Exception as e:
        log(f"\n[AVISO] Geracao do dashboard falhou: {e}")

    log("\n[CONCLUIDO]")


if __name__ == "__main__":
    asyncio.run(main())
