"""
Coleta semanal de métricas do Instagram.
Gera/atualiza três bases Excel:
  - base_seguidores.xlsx
  - base_interacoes_total.xlsx
  - base_interacoes_detalhadas.xlsx

Padrão: coleta a semana ISO anterior (segunda a domingo).
Quando rodado em uma segunda-feira, sempre captura a semana completa anterior.
"""

import json
import os
import requests
from datetime import datetime, timezone, timedelta
from calendar import monthrange
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")
BASE_URL = "https://graph.facebook.com/v21.0"
OUTPUT_DIR = os.path.dirname(__file__)


def load_config():
    with open(CONFIG_FILE, encoding="utf-8") as f:
        cfg = json.load(f)
    if os.environ.get("IG_ACCESS_TOKEN"):
        cfg["access_token"] = os.environ["IG_ACCESS_TOKEN"]
    return cfg


def api_get(path, params, silent=False):
    r = requests.get(f"{BASE_URL}/{path}", params=params, timeout=30)
    if not r.ok and not silent:
        try:
            print(f"    [API] {r.json().get('error', {}).get('message', r.text)}")
        except Exception:
            pass
    return r.json() if r.ok else {}


def periodo_atual():
    """
    Retorna (data_inicio, data_fim, iso_year, iso_week) para a coleta de hoje.
    data_inicio = primeiro dia do mês corrente
    data_fim    = hoje (a segunda-feira em que o script roda)
    """
    hoje = datetime.today().date()
    data_ini = hoje.replace(day=1)
    iso_year, iso_week, _ = hoje.isocalendar()
    return data_ini, hoje, iso_year, iso_week


def to_ts(date, end=False):
    """Converte date para timestamp UTC (inicio ou fim do dia)."""
    tz = timezone.utc
    if end:
        return int(datetime(date.year, date.month, date.day, 23, 59, 59, tzinfo=tz).timestamp())
    return int(datetime(date.year, date.month, date.day, 0, 0, 0, tzinfo=tz).timestamp())


# ─── COLETAS ────────────────────────────────────────────────────────────────

def get_seguidores_total(ig_id, token):
    data = api_get(ig_id, {"fields": "followers_count", "access_token": token})
    return data.get("followers_count")


def get_seguidores_ganhos(ig_id, token, since_ts, until_ts):
    """Novos seguidores no período via follower_count diário."""
    data = api_get(f"{ig_id}/insights", {
        "metric": "follower_count",
        "period": "day",
        "since": since_ts,
        "until": until_ts,
        "access_token": token,
    }, silent=True)
    ganhos = 0
    for item in data.get("data", []):
        if item.get("name") == "follower_count":
            for v in item.get("values", []):
                val = v.get("value", 0) or 0
                if val > 0:
                    ganhos += val
    return ganhos


def get_total_interacoes(ig_id, token, since_ts, until_ts):
    """total_interactions da semana via perfil/insights."""
    data = api_get(f"{ig_id}/insights", {
        "metric": "total_interactions",
        "period": "day",
        "metric_type": "total_value",
        "since": since_ts,
        "until": until_ts,
        "access_token": token,
    }, silent=True)
    for item in data.get("data", []):
        if item.get("name") == "total_interactions":
            val = item.get("total_value", {}).get("value")
            if val is not None:
                return val
    return 0


def get_interacoes_detalhadas(ig_id, token, since_ts, until_ts):
    """Curtidas, comentários, compartilhamentos e salvamentos da semana."""
    # Salvamentos via perfil (mais preciso)
    saves_data = api_get(f"{ig_id}/insights", {
        "metric": "saves",
        "period": "day",
        "metric_type": "total_value",
        "since": since_ts,
        "until": until_ts,
        "access_token": token,
    }, silent=True)
    salvamentos = 0
    for item in saves_data.get("data", []):
        if item.get("name") == "saves":
            val = item.get("total_value", {}).get("value")
            if val is not None:
                salvamentos = val

    # Curtidas, comentários e compartilhamentos via posts do período
    media_data = api_get(f"{ig_id}/media", {
        "fields": "id,timestamp,like_count,comments_count",
        "since": since_ts,
        "until": until_ts,
        "limit": 100,
        "access_token": token,
    })
    posts = media_data.get("data", [])
    curtidas = comentarios = compartilhamentos = 0
    for post in posts:
        curtidas += post.get("like_count", 0) or 0
        comentarios += post.get("comments_count", 0) or 0
        ins = api_get(f"{post['id']}/insights", {"metric": "shares", "access_token": token}, silent=True)
        for m in ins.get("data", []):
            if m.get("name") == "shares":
                val = (m.get("values") or [{}])[0].get("value") or m.get("value") or 0
                compartilhamentos += val

    return curtidas, comentarios, compartilhamentos, salvamentos, len(posts)


# ─── EXCEL ──────────────────────────────────────────────────────────────────

COR_HEADER = "1A3A5C"
COR_HEADER_FONT = "FFFFFF"
COR_ZEBRA = "EFF4FB"
COR_BORDA = "CCCCCC"

def estilo_header(cell):
    cell.font = Font(bold=True, color=COR_HEADER_FONT, size=11)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=COR_BORDA)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def estilo_dado(cell, zebra=False):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if zebra:
        cell.fill = PatternFill("solid", fgColor=COR_ZEBRA)
    thin = Side(style="thin", color=COR_BORDA)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def ajusta_colunas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


def carregar_ou_criar_wb(path, headers, descricao):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"
        ws.freeze_panes = "A2"
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            estilo_header(cell)
        ws.row_dimensions[1].height = 35
    return wb, ws


def linha_existe(ws, semana_label, perfil):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == semana_label and row[1] == perfil:
            return True
    return False


def salvar_wb(wb, ws, path):
    ajusta_colunas(ws)
    wb.save(path)


# ─── COLETA PRINCIPAL ───────────────────────────────────────────────────────

def coletar_semana(data_ini, data_fim, iso_year, iso_week, config):
    token = config["access_token"]
    profiles = config["profiles"]
    since_ts = to_ts(data_ini)
    until_ts = to_ts(data_fim, end=True)
    semana_label = f"S{iso_week:02d}/{iso_year}"
    data_coleta = datetime.now().strftime("%Y-%m-%d %H:%M")
    data_ini_str = data_ini.strftime("%d/%m/%Y")
    data_fim_str = data_fim.strftime("%d/%m/%Y")

    print(f"\nPeríodo: {data_ini_str} -> {data_fim_str}  |  {semana_label}\n")

    # ── Base 1: Seguidores ──────────────────────────────────────────────────
    path_seg = os.path.join(OUTPUT_DIR, "base_seguidores.xlsx")
    headers_seg = [
        "Semana", "Perfil", "Data inicio", "Data fim",
        "Seguidores total", "Seguidores ganhos", "Seguidores perdidos", "Crescimento liquido",
        "Coletado em"
    ]
    wb_seg, ws_seg = carregar_ou_criar_wb(path_seg, headers_seg, "Seguidores")

    # ── Base 2: Interações Total ────────────────────────────────────────────
    path_int = os.path.join(OUTPUT_DIR, "base_interacoes_total.xlsx")
    headers_int = [
        "Semana", "Perfil", "Data inicio", "Data fim",
        "Total interacoes",
        "Coletado em"
    ]
    wb_int, ws_int = carregar_ou_criar_wb(path_int, headers_int, "Interacoes Total")

    # ── Base 3: Interações Detalhadas ───────────────────────────────────────
    path_det = os.path.join(OUTPUT_DIR, "base_interacoes_detalhadas.xlsx")
    headers_det = [
        "Semana", "Perfil", "Data inicio", "Data fim",
        "Curtidas", "Comentarios", "Compartilhamentos", "Salvamentos",
        "Posts no periodo",
        "Coletado em"
    ]
    wb_det, ws_det = carregar_ou_criar_wb(path_det, headers_det, "Interacoes Detalhadas")

    for profile in profiles:
        name = profile["name"]
        ig_id = profile["ig_user_id"]

        if linha_existe(ws_seg, semana_label, name):
            print(f"  [{name}] {semana_label} já coletado, pulando (rode com --force para atualizar).")
            continue

        print(f"  [{name}] seguidores...", end=" ", flush=True)
        seg_total = get_seguidores_total(ig_id, token)

        print("ganhos...", end=" ", flush=True)
        seg_ganhos = get_seguidores_ganhos(ig_id, token, since_ts, until_ts)

        # Crescimento líquido: seguidores_total agora - seguidores_total do fim do mês anterior
        # Buscamos a última linha do mês anterior para esse perfil
        mes_atual_str = data_ini.strftime("%m/%Y")  # ex: "06/2026"
        prev_total = None
        for row in ws_seg.iter_rows(min_row=2, values_only=True):
            if row[1] == name and row[4] not in (None, ""):
                # row[3] = data_fim, ex "31/05/2026" — verifica se é mês anterior
                try:
                    row_mes = row[3][-7:]  # "MM/YYYY" do campo data_fim
                    if row_mes != mes_atual_str:
                        prev_total = row[4]
                except Exception:
                    pass
        crescimento_liq = ""
        seg_perdidos = ""
        try:
            if seg_total is not None and prev_total is not None:
                crescimento_liq = seg_total - int(prev_total)
                seg_perdidos = max(0, seg_ganhos - crescimento_liq)
        except Exception:
            pass

        print("interações...", end=" ", flush=True)
        total_int = get_total_interacoes(ig_id, token, since_ts, until_ts)

        print("detalhes...", end=" ", flush=True)
        curtidas, comentarios, compartilhamentos, salvamentos, n_posts = get_interacoes_detalhadas(
            ig_id, token, since_ts, until_ts
        )

        print("OK")

        n = ws_seg.max_row + 1
        zebra = n % 2 == 0
        dados_seg = [semana_label, name, data_ini_str, data_fim_str,
                     seg_total, seg_ganhos, seg_perdidos, crescimento_liq, data_coleta]
        for i, val in enumerate(dados_seg, 1):
            c = ws_seg.cell(row=n, column=i, value=val)
            estilo_dado(c, zebra)

        n2 = ws_int.max_row + 1
        zebra2 = n2 % 2 == 0
        dados_int = [semana_label, name, data_ini_str, data_fim_str, total_int, data_coleta]
        for i, val in enumerate(dados_int, 1):
            c = ws_int.cell(row=n2, column=i, value=val)
            estilo_dado(c, zebra2)

        n3 = ws_det.max_row + 1
        zebra3 = n3 % 2 == 0
        dados_det = [semana_label, name, data_ini_str, data_fim_str,
                     curtidas, comentarios, compartilhamentos, salvamentos, n_posts, data_coleta]
        for i, val in enumerate(dados_det, 1):
            c = ws_det.cell(row=n3, column=i, value=val)
            estilo_dado(c, zebra3)

    salvar_wb(wb_seg, ws_seg, path_seg)
    salvar_wb(wb_int, ws_int, path_int)
    salvar_wb(wb_det, ws_det, path_det)

    print(f"\n  Salvo:")
    print(f"  {path_seg}")
    print(f"  {path_int}")
    print(f"  {path_det}\n")


# ─── ENTRADA ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--semana", type=int, help="Numero da semana ISO (ex: 25)")
    parser.add_argument("--ano", type=int, help="Ano (ex: 2026)")
    args = parser.parse_args()

    cfg = load_config()

    if args.semana and args.ano:
        # Semana específica: segunda-feira daquela semana ISO como data_fim, inicio do mês como data_ini
        import datetime as dt
        segunda = dt.date.fromisocalendar(args.ano, args.semana, 1)
        data_ini = segunda.replace(day=1)
        data_fim = segunda
        iso_year, iso_week = args.ano, args.semana
    else:
        data_ini, data_fim, iso_year, iso_week = periodo_atual()

    coletar_semana(data_ini, data_fim, iso_year, iso_week, cfg)
